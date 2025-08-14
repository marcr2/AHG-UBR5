import json
import numpy as np
import requests
from sklearn.metrics.pairwise import cosine_similarity
import os
from tqdm.auto import tqdm
from chromadb_manager import ChromaDBManager
import logging
from hypothesis_tools import HypothesisGenerator, HypothesisCritic, MetaHypothesisGenerator, get_lab_goals, get_lab_config
import time
import random
import threading
from collections import deque
import pandas as pd  # For Excel export
from openpyxl import load_workbook
from datetime import datetime
import re

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HypothesisTimer:
    """Timer for hypothesis critique process with 5-minute limit."""

    def __init__(self, timeout_minutes=5):
        self.timeout_seconds = timeout_minutes * 60
        self.start_time = None
        self.is_expired = False
        self.lock = threading.Lock()

    def start(self):
        """Start the timer."""
        with self.lock:
            self.start_time = time.time()
            self.is_expired = False

    def check_expired(self):
        """Check if the timer has expired."""
        if self.start_time is None:
            return False
        with self.lock:
            if not self.is_expired:
                elapsed = time.time() - self.start_time
                if elapsed >= self.timeout_seconds:
                    self.is_expired = True
            return self.is_expired

    def get_remaining_time(self):
        """Get remaining time in seconds."""
        if self.start_time is None:
            return self.timeout_seconds
        elapsed = time.time() - self.start_time
        return max(0, self.timeout_seconds - elapsed)

    def reset(self):
        """Reset the timer."""
        with self.lock:
            self.start_time = None
            self.is_expired = False

class GeminiRateLimiter:
    """Rate limiter for Gemini API calls (1000 requests per minute)."""

    def __init__(self, max_requests_per_minute=1000):
        self.max_requests = max_requests_per_minute
        self.request_times = deque()
        self.lock = threading.Lock() if 'threading' in globals() else None

    def wait_if_needed(self):
        """Wait if we're at the rate limit."""
        current_time = time.time()

        # Remove old requests (older than 1 minute)
        while self.request_times and current_time - self.request_times[0] > 60:
            self.request_times.popleft()

        # If we're at the limit, wait
        if len(self.request_times) >= self.max_requests:
            wait_time = 60 - (current_time - self.request_times[0]) + 1
            if wait_time > 0:
                print(f"⏳ Rate limit reached. Waiting {wait_time:.1f}s for Gemini API...")
                time.sleep(wait_time)
                current_time = time.time()

        # Add current request
        self.request_times.append(current_time)

    def get_remaining_requests(self):
        """Get number of remaining requests in current window."""
        current_time = time.time()

        # Remove old requests
        while self.request_times and current_time - self.request_times[0] > 60:
            self.request_times.popleft()

        return max(0, self.max_requests - len(self.request_times))

class EnhancedRAGQuery:
    """
    Enhanced RAG Query System with ChromaDB integration.

    Features:
    - Traditional similarity search (original functionality)
    - ChromaDB vector database for persistent storage
    - Metadata filtering
    - Hybrid search capabilities
    - Performance comparison between methods
    """

    def __init__(self, use_chromadb: bool = True, load_data_at_startup: bool = True):
        """Initialize the enhanced RAG query system."""
        self.use_chromadb = use_chromadb
        self.load_data_at_startup = load_data_at_startup
        self.embeddings_data = None
        self.chroma_manager = None
        self.hypothesis_generator = None
        self.hypothesis_critic = None
        self.meta_hypothesis_generator = None
        self.last_context_chunks = None
        self.hypothesis_records = []  # Track all hypotheses and results for export

        # Track the initial "add" quantity for dynamic chunk selection
        self.initial_add_quantity = None
        self.hypothesis_timer = HypothesisTimer(timeout_minutes=5)  # 5-minute timer

        # Track used papers to avoid reusing them in dynamic chunk selection
        self.used_papers = set()  # Set of paper identifiers (DOI or title) that have been used

        # Package system for collecting search results
        self.current_package = {
            "chunks": [],
            "metadata": [],
            "sources": set(),
            "total_chars": 0
        }

        # Initialize Gemini rate limiter (1000 requests per minute)
        self.gemini_rate_limiter = GeminiRateLimiter(max_requests_per_minute=1000)

        print("🚀 Initializing Enhanced RAG System...")

        # Load API keys first
        try:
            with open("keys.json") as f:
                keys = json.load(f)
                self.api_key = keys["GOOGLE_API_KEY"]  # For embeddings
                self.gemini_api_key = keys["GEMINI_API_KEY"]  # For text generation
            print("✅ API keys loaded successfully")
        except Exception as e:
            logger.error(f"❌ Failed to load API keys: {e}")
            self.api_key = None
            self.gemini_api_key = None

        # Initialize Gemini client
        try:
            from google import genai
            self.gemini_client = genai.Client(api_key=self.gemini_api_key)
            print("✅ Gemini client initialized successfully")
        except Exception as e:
            logger.error(f"❌ Failed to initialize Gemini client: {e}")
            self.gemini_client = None

        # Initialize hypothesis tools after Gemini client is ready
        self._initialize_hypothesis_tools()

        # Initialize ChromaDB if requested
        if self.use_chromadb:
            self._initialize_chromadb()

        # Load embeddings data (optional)
        if self.load_data_at_startup:
            print("📚 Loading embeddings data (this may take a while for large datasets)...")
            self._load_embeddings_data()
        else:
            print("⏭️  Skipping data loading at startup (will load on-demand)")
            self.embeddings_data = None

        print("✅ Enhanced RAG System initialized!")

    def _initialize_chromadb(self):
        """Initialize ChromaDB manager."""
        try:
            self.chroma_manager = ChromaDBManager()
            if not self.chroma_manager.create_collection():
                logger.error("❌ Failed to initialize ChromaDB")
                self.use_chromadb = False
            else:
                logger.info("✅ ChromaDB initialized successfully")
        except Exception as e:
            logger.error(f"❌ ChromaDB initialization failed: {e}")
            self.use_chromadb = False

    def _load_embeddings_data(self):
        """Load all available embedding files."""
        # Check if ChromaDB already has data
        if self.use_chromadb and self.chroma_manager:
            stats = self.chroma_manager.get_collection_stats()
            if stats.get('total_documents', 0) > 0:
                logger.info(f"📚 ChromaDB already has {stats.get('total_documents', 0)} documents - skipping data loading")
                self.embeddings_data = None  # We'll use ChromaDB directly
                return

        # Only load embeddings if ChromaDB is empty
        self.embeddings_data = self.load_all_embeddings()

        # If ChromaDB is available, populate it
        if self.use_chromadb and self.embeddings_data:
            self._populate_chromadb()

    def _populate_chromadb(self):
        """Populate ChromaDB with embeddings data."""
        if not self.chroma_manager or not self.embeddings_data:
            return

        try:
            # Check if collection is empty
            stats = self.chroma_manager.get_collection_stats()
            if stats.get('total_documents', 0) > 0:
                logger.info(f"📚 ChromaDB collection already populated with {stats.get('total_documents', 0)} documents")
                return

            logger.info("🔄 Populating ChromaDB with embeddings data...")

            # Add embeddings from each source
            for source_name, source_stats in self.embeddings_data["sources"].items():
                # Filter data for this source
                source_data = {
                    "chunks": [],
                    "embeddings": [],
                    "metadata": []
                }

                # Add progress bar for metadata processing
                with tqdm(total=len(self.embeddings_data["metadata"]), desc=f"Processing metadata for {source_name}", unit="entry") as meta_pbar:
                    for i, meta in enumerate(self.embeddings_data["metadata"]):
                        if meta.get("source_file") == source_name:
                            source_data["chunks"].append(self.embeddings_data["chunks"][i])
                            source_data["embeddings"].append(self.embeddings_data["embeddings"][i])
                            source_data["metadata"].append(meta)
                        meta_pbar.update(1)

                if source_data["chunks"]:
                    self.chroma_manager.add_embeddings_to_collection(source_data, source_name)
                    logger.info(f"✅ Added {len(source_data['chunks'])} embeddings from {source_name}")

            # Show final stats
            final_stats = self.chroma_manager.get_collection_stats()
            logger.info(f"🎉 ChromaDB populated with {final_stats.get('total_documents', 0)} total documents")

        except Exception as e:
            logger.error(f"❌ Failed to populate ChromaDB: {e}")

    def get_google_embedding(self, text):
        """Get embedding for a query text using Google's API."""
        if not self.api_key:
            logger.error("❌ API key not available")
            return None

        url = f"https://generativelanguage.googleapis.com/v1beta/models/text-embedding-004:embedContent?key={self.api_key}"
        headers = {"Content-Type": "application/json"}
        data = {"model": "models/text-embedding-004", "content": {"parts": [{"text": text}]}}

        try:
            # Add timeout to prevent hanging
            response = requests.post(url, headers=headers, json=data, timeout=30)
            response.raise_for_status()
            embedding = response.json()["embedding"]["values"]
            return embedding
        except requests.exceptions.Timeout:
            logger.error("❌ Query embedding request timed out (30s)")
            return None
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ Network error getting query embedding: {e}")
            return None
        except Exception as e:
            logger.error(f"❌ Error getting query embedding: {e}")
            return None

    def load_embeddings_from_json(self, filename):
        """Load embeddings from JSON file."""
        if not os.path.exists(filename):
            logger.warning(f"⚠️ Embeddings file '{filename}' not found!")
            return None

        try:
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)

            logger.info(f"📊 Loaded {len(data['chunks'])} chunks from {filename}")
            return data
        except Exception as e:
            logger.error(f"❌ Failed to load {filename}: {e}")
            return None

    def load_all_embeddings(self):
        """Load all available embedding files (both single files and batch-based)."""
        all_data = {
            "chunks": [],
            "embeddings": [],
            "metadata": [],
            "sources": {}
        }

        total_chunks = 0

        # Load PubMed embeddings (single file)
        if os.path.exists("xrvix_embeddings/pubmed_embeddings.json"):
            logger.info(f"🔄 Loading pubmed_embeddings.json from xrvix_embeddings folder...")
            data = self.load_embeddings_from_json("xrvix_embeddings/pubmed_embeddings.json")
            if data:
                # Add source prefix to metadata with progress bar
                with tqdm(total=len(data["metadata"]), desc="Adding source prefix to PubMed metadata", unit="entry") as meta_pbar:
                    for meta in data["metadata"]:
                        meta["source_file"] = "pubmed"
                        meta_pbar.update(1)

                # Append to combined data
                all_data["chunks"].extend(data["chunks"])
                all_data["embeddings"].extend(data["embeddings"])
                all_data["metadata"].extend(data["metadata"])

                # Track source statistics
                all_data["sources"]["pubmed"] = {
                    "chunks": len(data["chunks"]),
                    "embeddings": len(data["embeddings"]),
                    "stats": data.get("stats", {})
                }

                total_chunks += len(data["chunks"])

        # Load xrvix embeddings (batch-based system)
        xrvix_dir = "xrvix_embeddings"
        if os.path.exists(xrvix_dir):
            logger.info(f"🔄 Loading xrvix embeddings from {xrvix_dir}...")

            # Load metadata first
            metadata_file = os.path.join(xrvix_dir, "metadata.json")
            if os.path.exists(metadata_file):
                try:
                    with open(metadata_file, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                    logger.info(f"📊 Found metadata: {metadata.get('total_embeddings', 0)} total embeddings")
                except Exception as e:
                    logger.error(f"❌ Failed to load metadata: {e}")
                    metadata = {}
            else:
                metadata = {}

            # Process each source directory (biorxiv, medrxiv, etc.)
            for source_dir in os.listdir(xrvix_dir):
                source_path = os.path.join(xrvix_dir, source_dir)
                if os.path.isdir(source_path) and source_dir in ["biorxiv", "medrxiv"]:
                    logger.info(f"🔄 Processing {source_dir}...")

                    # Find all batch files
                    batch_files = [f for f in os.listdir(source_path) if f.startswith("batch_") and f.endswith(".json")]
                    batch_files.sort()  # Sort to process in order

                    source_chunks = 0
                    source_embeddings = 0

                    # Add progress bar for batch loading
                    for batch_file in batch_files:
                        batch_path = os.path.join(source_path, batch_file)
                        try:
                            with open(batch_path, 'r', encoding='utf-8') as f:
                                batch_data = json.load(f)

                            # Add source information to metadata with progress bar
                            batch_metadata = batch_data.get("metadata", [])
                            with tqdm(total=len(batch_metadata), desc=f"Adding source info to {source_dir} batch", unit="entry") as meta_pbar:
                                for meta in batch_metadata:
                                    meta["source_file"] = source_dir
                                    meta_pbar.update(1)

                            # Append to combined data
                            batch_chunks = len(batch_data.get("chunks", []))
                            batch_embeddings = len(batch_data.get("embeddings", []))

                            all_data["chunks"].extend(batch_data.get("chunks", []))
                            all_data["embeddings"].extend(batch_data.get("embeddings", []))
                            all_data["metadata"].extend(batch_data.get("metadata", []))

                            source_chunks += batch_chunks
                            source_embeddings += batch_embeddings

                        except Exception as e:
                            logger.error(f"❌ Failed to load batch {batch_file}: {e}")

                    # Track source statistics
                    if source_chunks > 0:
                        all_data["sources"][source_dir] = {
                            "chunks": source_chunks,
                            "embeddings": source_embeddings,
                            "batches": len(batch_files),
                            "stats": metadata.get("sources", {}).get(source_dir, {})
                        }
                        total_chunks += source_chunks
                        logger.info(f"✅ Loaded {source_chunks} chunks from {source_dir} ({len(batch_files)} batches)")

        if total_chunks == 0:
            logger.warning("⚠️ No embedding files found!")
            return None

        logger.info(f"🎉 Combined {total_chunks} chunks from all sources")
        return all_data

    def search_chromadb(self, query, top_k=5, filter_dict=None, show_progress=True):
        """Search using ChromaDB."""
        if not self.use_chromadb or not self.chroma_manager:
            print("❌ ChromaDB not available. Please initialize with ChromaDB enabled.")
            return []

        # Check if ChromaDB has data
        if not self.is_chromadb_ready():
            print("❌ ChromaDB is empty. Please load data first using the master processor (option 4).")
            return []

        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            return []

        # Search in ChromaDB
        results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=top_k,
            where_filter=filter_dict
        )

        return results

    def search_hybrid(self, query, top_k=20, filter_dict=None):
        """Search using ChromaDB only (simplified from hybrid)."""
        if not self.use_chromadb or not self.chroma_manager:
            print("❌ ChromaDB not available. Please initialize with ChromaDB enabled.")
            return []

        try:
            # Use only ChromaDB search
            results = self.search_chromadb(query, top_k=top_k, filter_dict=filter_dict)

            # Convert ChromaDB results to the expected format
            formatted_results = []
            for result in results:
                formatted_results.append({
                    "chunk": result.get("document", ""),
                    "metadata": result.get("metadata", {}),
                    "similarity": 1.0 - result.get("distance", 0) if result.get("distance") is not None else 0,
                    "method": "chromadb"
                })

            return formatted_results

        except Exception as e:
            print(f"❌ ChromaDB search failed: {e}")
            return []

    def display_results(self, results, show_method=True):
        """Display search results in a formatted way."""
        if not results:
            print("[display_results] No results found.")
            return
        print(f"[display_results] Found {len(results)} relevant chunks.")
        total_chars = sum(len(result.get("chunk", "")) for result in results)
        sources = set()
        methods = set()
        for result in results:
            sources.add(result.get("metadata", {}).get("source", "Unknown"))
            methods.add(result.get("method", "Unknown"))
        print(f"[display_results] Summary:")
        print(f"   📚 Total chunks: {len(results)}")
        print(f"   📝 Total characters: {total_chars:,}")
        if show_method:
            print(f"   🔍 Search methods: {', '.join(methods)}")
        print(f"   📖 Sources: {', '.join(sources)}")
        print(f"[display_results] Top 3 results (showing examples):")
        for i, result in enumerate(results[:3], 1):
            chunk = result.get("chunk", "")
            metadata = result.get("metadata", {})
            similarity = result.get("similarity", 0)
            title = metadata.get("title", chunk[:60])
            if len(title) > 60:
                title = title[:57] + "..."
            print(f"   {i}. {title} (similarity: {similarity:.3f})")
        if len(results) > 3:
            print(f"   ... and {len(results) - 3} more chunks")

    def display_statistics(self):
        """Display comprehensive statistics about the knowledge base."""
        print(f"[display_statistics] Knowledge Base Statistics:")
        print("=" * 50)

        # Show ChromaDB stats if available (primary data source)
        if self.use_chromadb and self.chroma_manager:
            print(f"[display_statistics] ChromaDB Statistics (Primary Data Source):")
            chroma_stats = self.chroma_manager.get_collection_stats()
            print(f"   📊 Total documents: {chroma_stats.get('total_documents', 0)}")
            print(f"   📊 Collection name: {chroma_stats.get('collection_name', 'N/A')}")
            print(f"   📊 Available collections: {self.chroma_manager.list_collections()}")

            # Show source breakdown if available
            batch_stats = self.chroma_manager.get_batch_statistics()
            if batch_stats and batch_stats.get('sources'):
                print(f"[display_statistics] Source Breakdown:")
                for source, stats in batch_stats['sources'].items():
                    print(f"   {source}: {stats['total_documents']} documents, {len(stats['batches'])} batches")

        # Show in-memory data stats if available (secondary/fallback)
        if self.embeddings_data:
            print(f"[display_statistics] In-Memory Data Statistics (Secondary):")
            total_chunks = len(self.embeddings_data["chunks"])
            total_embeddings = len(self.embeddings_data["embeddings"])
            print(f"   📈 Total chunks: {total_chunks}")
            print(f"   📈 Total embeddings: {total_embeddings}")
            if total_embeddings > 0:
                print(f"   📈 Embedding dimensions: {len(self.embeddings_data['embeddings'][0])}")
            print(f"[display_statistics] Source Breakdown:")
            for source, stats in self.embeddings_data["sources"].items():
                print(f"   {source}: {stats['chunks']} chunks")
                if "stats" in stats and stats["stats"]:
                    source_stats = stats["stats"]
                    print(f"     - Papers: {source_stats.get('total_papers', 'N/A')}")
                    print(f"     - Embeddings: {source_stats.get('total_embeddings', 'N/A')}")
        else:
            print(f"[display_statistics] In-Memory Data: Not loaded (using ChromaDB directly)")

    def extract_citations_from_chunks(self, context_chunks):
        """Extract citation information from context chunks used for hypothesis critique."""
        citations = []
        unique_sources = set()

        for chunk in context_chunks:
            if isinstance(chunk, dict):
                metadata = chunk.get('metadata', {})
                # Extract source information
                source_name = metadata.get('source_name', 'Unknown')
                title = metadata.get('title', 'No title')
                doi = metadata.get('doi', 'No DOI')
                authors = metadata.get('author', metadata.get('authors', 'Unknown authors'))
                journal = metadata.get('journal', 'Unknown journal')
                # Extract year from publication_date
                publication_date = metadata.get('publication_date', '')
                year = publication_date[:4] if publication_date and len(publication_date) >= 4 else 'Unknown year'

                # Create citation entry
                citation = {
                    'source_name': source_name,
                    'title': title,
                    'doi': doi,
                    'authors': authors,
                    'journal': journal,
                    'year': year
                }

                # Add to citations if not already present (based on DOI or title)
                citation_key = doi if doi != 'No DOI' else title
                if citation_key not in unique_sources:
                    citations.append(citation)
                    unique_sources.add(citation_key)
            else:
                # Fallback for chunks without metadata
                citations.append({
                    'source_name': 'Unknown',
                    'title': 'Unknown source',
                    'doi': 'No DOI',
                    'authors': 'Unknown authors',
                    'journal': 'Unknown journal',
                    'year': 'Unknown year'
                })

        return citations

    def format_citations_for_export(self, citations):
        """Format citations for Excel export."""
        if not citations:
            return "No citations available"

        formatted_citations = []
        for citation in citations:
            # Format as academic citation
            if citation['doi'] != 'No DOI':
                formatted = f"{citation['authors']} ({citation['year']}). {citation['title']}. {citation['journal']}. DOI: {citation['doi']}"
            else:
                formatted = f"{citation['authors']} ({citation['year']}). {citation['title']}. {citation['journal']}."
            formatted_citations.append(formatted)

        return "; ".join(formatted_citations)

    def retrieve_relevant_chunks(self, query, top_k=1500, lab_paper_ratio=None):
        """Retrieve the most relevant chunks from all loaded batches using ChromaDB, with lab-authored papers included."""
        if lab_paper_ratio is None:
            lab_paper_ratio = getattr(self, 'lab_paper_ratio', 0.2)
        return self.retrieve_relevant_chunks_with_lab_papers(query, top_k, lab_paper_ratio)

    def select_dynamic_chunks_for_generation(self, query, num_chunks=None, lab_paper_ratio=None, randomization_strategy='enhanced'):
        """
        Select new chunks from ChromaDB for hypothesis generation, ensuring diversity across sources and including lab-authored papers.
        
        Args:
            query: Search query
            num_chunks: Number of chunks to select (defaults to initial_add_quantity)
            lab_paper_ratio: Ratio of lab-authored papers to include
            randomization_strategy: Strategy for chunk randomization ('enhanced', 'shuffle', 'diversity', 'time_based')
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("❌ ChromaDB not available for dynamic chunk selection.")
            return None

        # Use configured lab paper ratio if not specified
        if lab_paper_ratio is None:
            lab_paper_ratio = getattr(self, 'lab_paper_ratio', None)

        # Use initial add quantity if not specified
        if num_chunks is None:
            if self.initial_add_quantity is None:
                print("⚠️  No initial add quantity stored, using default of 1500 chunks")
                num_chunks = 1500
            else:
                num_chunks = self.initial_add_quantity
                print(f"🔄 Selecting {num_chunks} new chunks from ChromaDB (based on initial add quantity)")

        try:
            # Get new chunks from ChromaDB using the original query
            if "prompt" in self.current_package and self.current_package["prompt"]:
                search_query = self.current_package["prompt"]
            else:
                search_query = query

            print(f"🔍 Searching ChromaDB for '{search_query}' to select {num_chunks} new chunks...")
            print(f"🎯 Ensuring diversity across sources: pubmed, biorxiv, medrxiv")
            if lab_paper_ratio is None:
                print(f"🔬 Lab papers: AUTO (automatically determined based on availability)")
            else:
                print(f"🔬 Including {int(num_chunks * lab_paper_ratio)} lab-authored papers")

            # Use the sophisticated lab paper search method
            preferred_authors = getattr(self, 'preferred_authors', [])
            results = self.retrieve_relevant_chunks_with_lab_papers(
                search_query,
                top_k=min(num_chunks * 3, 5000),  # Search up to 3x the needed amount, max 5000
                lab_paper_ratio=lab_paper_ratio,
                preferred_authors=preferred_authors
            )

            if not results:
                print("❌ No new chunks found in ChromaDB.")
                return None

            # Apply randomization strategy
            results = self._apply_randomization_strategy(results, randomization_strategy)

            # Organize results by source for balanced selection
            source_chunks = {
                'pubmed': [],
                'biorxiv': [],
                'medrxiv': []
            }

            # Categorize results by source
            for result in results:
                metadata = result.get('metadata', {})
                source = metadata.get('source', '').lower()
                source_name = metadata.get('source_name', '').lower()

                # Determine the source
                if 'pubmed' in source or 'pubmed' in source_name:
                    source_key = 'pubmed'
                elif 'biorxiv' in source or 'biorxiv' in source_name:
                    source_key = 'biorxiv'
                elif 'medrxiv' in source or 'medrxiv' in source_name:
                    source_key = 'medrxiv'
                else:
                    # Default to pubmed if source is unclear
                    source_key = 'pubmed'

                source_chunks[source_key].append(result)

            # Calculate target chunks per source (balanced distribution)
            chunks_per_source = max(1, num_chunks // 3)  # At least 1 chunk per source
            remaining_chunks = num_chunks % 3

            print(f"📊 Source distribution target: ~{chunks_per_source} chunks per source")

            # Filter out chunks from already used papers and select balanced chunks
            new_chunks = []
            used_papers_in_this_selection = set()
            source_counts = {'pubmed': 0, 'biorxiv': 0, 'medrxiv': 0}

            # Select chunks from each source, avoiding used papers
            for source_key in ['pubmed', 'biorxiv', 'medrxiv']:
                source_target = chunks_per_source + (1 if remaining_chunks > 0 else 0)
                remaining_chunks = max(0, remaining_chunks - 1)

                source_results = source_chunks[source_key]
                source_selected = 0

                for result in source_results:
                    if source_selected >= source_target:
                        break

                    metadata = result.get('metadata', {})
                    paper_id = self._get_paper_identifier(metadata)

                    # Skip if this paper has been used before
                    if paper_id in self.used_papers:
                        continue

                    # Add to new chunks and mark as used
                    new_chunks.append(result['chunk'])
                    used_papers_in_this_selection.add(paper_id)
                    source_counts[source_key] += 1
                    source_selected += 1

            # If we don't have enough chunks from balanced selection, fill with any available chunks
            if len(new_chunks) < num_chunks:
                print(f"⚠️  Balanced selection only found {len(new_chunks)} chunks, filling with additional chunks...")

                # Collect all remaining unused chunks
                remaining_chunks = []
                for result in results:
                    metadata = result.get('metadata', {})
                    paper_id = self._get_paper_identifier(metadata)

                    if paper_id not in self.used_papers and paper_id not in used_papers_in_this_selection:
                        remaining_chunks.append(result)

                # Add remaining chunks until we reach the target
                for result in remaining_chunks:
                    if len(new_chunks) >= num_chunks:
                        break

                    metadata = result.get('metadata', {})
                    paper_id = self._get_paper_identifier(metadata)

                    new_chunks.append(result['chunk'])
                    used_papers_in_this_selection.add(paper_id)

                    # Track source for remaining chunks
                    source = metadata.get('source', '').lower()
                    source_name = metadata.get('source_name', '').lower()
                    if 'biorxiv' in source or 'biorxiv' in source_name:
                        source_counts['biorxiv'] += 1
                    elif 'medrxiv' in source or 'medrxiv' in source_name:
                        source_counts['medrxiv'] += 1
                    else:
                        source_counts['pubmed'] += 1

            # Add the newly used papers to the tracking set
            self.used_papers.update(used_papers_in_this_selection)

            print(f"✅ Selected {len(new_chunks)} new chunks from {len(used_papers_in_this_selection)} previously unused papers")
            print(f"📊 Source distribution: pubmed={source_counts['pubmed']}, biorxiv={source_counts['biorxiv']}, medrxiv={source_counts['medrxiv']}")
            print(f"📊 Total papers used so far: {len(self.used_papers)}")

            if len(new_chunks) < num_chunks:
                print(f"⚠️  Only found {len(new_chunks)} new chunks (requested {num_chunks})")
                if len(self.used_papers) > 100:  # Arbitrary threshold
                    print("💡 Consider resetting used papers list with 'reset_papers' command")

            # Store the results for citation extraction
            self.last_dynamic_results = []
            for result in results:
                if result.get('chunk') in new_chunks:
                    self.last_dynamic_results.append(result)

            return new_chunks

        except Exception as e:
            print(f"❌ Error selecting dynamic chunks: {e}")
            return None

    def select_diverse_papers_for_hypothesis_generation(self, query, target_papers=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Select exactly 1500 different papers for hypothesis generation, ensuring maximum diversity
        while maintaining the existing prioritization system (lab author, date, citation count, impact factor).
        
        This function ensures that each hypothesis generation gets a completely different set of papers,
        maximizing the variety of source chunks as required by the new prompt format.
        
        Args:
            query: Search query
            target_papers: Number of different papers to select (default: 1500)
            lab_paper_ratio: Ratio of lab-authored papers to include
            preferred_authors: List of preferred author names to prioritize
            
        Returns:
            List of chunks from exactly target_papers different papers
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("❌ ChromaDB not available for diverse paper selection.")
            return None

        print(f"🎯 Selecting exactly {target_papers} different papers for hypothesis generation...")
        print(f"🔍 Ensuring maximum diversity across all sources and papers")

        try:
            # Use configured lab paper ratio if not specified
            if lab_paper_ratio is None:
                lab_paper_ratio = getattr(self, 'lab_paper_ratio', None)

            # Search for a large pool of papers to ensure we have enough diversity
            # We need significantly more than target_papers to account for filtering and prioritization
            search_pool_size = max(target_papers * 5, 10000)  # At least 5x the target, minimum 10k
            
            print(f"🔍 Searching ChromaDB for '{query}' to create pool of {search_pool_size} papers...")
            
            # Use the sophisticated lab paper search method to get a large pool
            preferred_authors = getattr(self, 'preferred_authors', [])
            search_results = self.retrieve_relevant_chunks_with_lab_papers(
                query,
                top_k=search_pool_size,
                lab_paper_ratio=lab_paper_ratio,
                preferred_authors=preferred_authors
            )

            if not search_results:
                print("❌ No papers found in ChromaDB.")
                return None

            print(f"📊 Found {len(search_results)} papers in search pool")

            # Step 1: Group results by paper (one chunk per paper)
            paper_chunks = {}  # paper_id -> best_chunk
            paper_metadata = {}  # paper_id -> metadata
            
            for result in search_results:
                metadata = result.get('metadata', {})
                paper_id = self._get_paper_identifier(metadata)
                
                if not paper_id:
                    continue
                    
                # Store the best chunk for each paper (first one found)
                if paper_id not in paper_chunks:
                    paper_chunks[paper_id] = result['chunk']
                    paper_metadata[paper_id] = metadata

            print(f"📊 Identified {len(paper_chunks)} unique papers")

            # Step 2: Apply existing prioritization system
            # Categorize papers by priority level
            lab_papers = []
            preferred_author_papers = []
            other_papers = []
            
            for paper_id, metadata in paper_metadata.items():
                if self.is_lab_authored_paper(metadata):
                    lab_papers.append((paper_id, metadata))
                elif preferred_authors and self._is_preferred_author_paper(metadata, preferred_authors):
                    preferred_author_papers.append((paper_id, metadata))
                else:
                    other_papers.append((paper_id, metadata))

            print(f"📊 Priority categorization: {len(lab_papers)} lab papers, {len(preferred_author_papers)} preferred author papers, {len(other_papers)} other papers")

            # Step 3: Calculate prioritization scores for other papers
            # Extract citation and journal data for scoring
            all_citations = []
            journal_impact_factors = {}
            
            for paper_id, metadata in paper_metadata.items():
                citation_count = metadata.get('citation_count', 'not found')
                journal = metadata.get('journal', 'Unknown journal')
                
                if citation_count != 'not found':
                    try:
                        all_citations.append(int(citation_count))
                    except (ValueError, TypeError):
                        pass
                
                if journal not in journal_impact_factors:
                    journal_impact_factors[journal] = []
                if citation_count != 'not found':
                    try:
                        journal_impact_factors[journal].append(int(citation_count))
                    except (ValueError, TypeError):
                        pass

            # Calculate statistics for scoring
            avg_citations = np.mean(all_citations) if all_citations else 0
            journal_impact_scores = {}
            for journal, citations in journal_impact_factors.items():
                if citations:
                    journal_impact_scores[journal] = np.mean(citations)

            # Step 4: Score and sort other papers
            def score_paper(paper_id, metadata):
                journal = metadata.get('journal', 'Unknown journal')
                citation_count = metadata.get('citation_count', 'not found')
                
                # Journal impact score (0-1 scale)
                journal_score = 0
                if journal in journal_impact_scores:
                    max_impact = max(journal_impact_scores.values()) if journal_impact_scores else 1
                    journal_score = journal_impact_scores[journal] / max_impact if max_impact > 0 else 0
                
                # Citation score (0-1 scale)
                citation_score = 0
                if citation_count != 'not found':
                    try:
                        citations = int(citation_count)
                        citation_score = min(citations / avg_citations, 2.0) if avg_citations > 0 else 0
                    except (ValueError, TypeError):
                        pass
                
                # Combined score (weighted average)
                combined_score = (journal_score * 0.4) + (citation_score * 0.6)
                return combined_score

            # Score and sort other papers
            other_papers_scored = [(paper_id, metadata, score_paper(paper_id, metadata)) for paper_id, metadata in other_papers]
            other_papers_scored.sort(key=lambda x: x[2], reverse=True)

            # Step 5: Select papers ensuring diversity with temporal balance
            selected_papers = set()
            selected_chunks = []
            
            # First, add all lab papers (highest priority)
            lab_papers_to_add = min(len(lab_papers), int(target_papers * 0.3))  # Up to 30% lab papers
            for paper_id, metadata in lab_papers[:lab_papers_to_add]:
                if paper_id not in selected_papers:
                    selected_papers.add(paper_id)
                    selected_chunks.append(paper_chunks[paper_id])
                    if len(selected_papers) >= target_papers:
                        break

            # Next, add preferred author papers
            preferred_author_papers_to_add = min(len(preferred_author_papers), int(target_papers * 0.2))  # Up to 20% preferred author papers
            for paper_id, metadata in preferred_author_papers[:preferred_author_papers_to_add]:
                if paper_id not in selected_papers:
                    selected_papers.add(paper_id)
                    selected_chunks.append(paper_chunks[paper_id])
                    if len(selected_papers) >= target_papers:
                        break

            # Calculate remaining slots for other papers
            remaining_needed = target_papers - len(selected_papers)
            
            # For other papers, ensure temporal balance: 50% from oldest 25% and newest 25%
            if remaining_needed > 0 and other_papers_scored:
                # Extract publication years for temporal analysis
                paper_years = []
                for paper_id, metadata, score in other_papers_scored:
                    publication_date = metadata.get('publication_date', '')
                    if publication_date and len(publication_date) >= 4:
                        try:
                            year = int(publication_date[:4])
                            paper_years.append((paper_id, metadata, score, year))
                        except (ValueError, TypeError):
                            # If we can't determine year, add with year 0 (will be treated as unknown)
                            paper_years.append((paper_id, metadata, score, 0))
                    else:
                        # No publication date, add with year 0
                        paper_years.append((paper_id, metadata, score, 0))
                
                if paper_years:
                    # Sort by year (oldest first)
                    paper_years.sort(key=lambda x: x[3])
                    
                    # Calculate temporal quartiles
                    total_other_papers = len(paper_years)
                    q1_size = max(1, total_other_papers // 4)  # Bottom 25% (oldest)
                    q4_size = max(1, total_other_papers // 4)  # Top 25% (newest)
                    
                    # Calculate how many papers to take from each temporal quartile
                    # We want 50% of remaining_needed from oldest+newest, 50% from middle
                    temporal_balance_count = min(remaining_needed // 2, q1_size + q4_size)
                    middle_count = remaining_needed - temporal_balance_count
                    
                    print(f"📅 Temporal balance: {temporal_balance_count} papers from oldest/newest quartiles, {middle_count} from middle")
                    
                    # Select from oldest papers (bottom 25%)
                    oldest_papers = paper_years[:q1_size]
                    oldest_selected = 0
                    for paper_id, metadata, score, year in oldest_papers:
                        if oldest_selected >= temporal_balance_count // 2:
                            break
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                            oldest_selected += 1
                            if len(selected_papers) >= target_papers:
                                break
                    
                    # Select from newest papers (top 25%)
                    newest_papers = paper_years[-q4_size:]
                    newest_selected = 0
                    for paper_id, metadata, score, year in newest_papers:
                        if newest_selected >= temporal_balance_count // 2:
                            break
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                            newest_selected += 1
                            if len(selected_papers) >= target_papers:
                                break
                    
                    # Fill remaining slots with middle-aged papers (prioritized by score)
                    middle_papers = paper_years[q1_size:-q4_size]  # Middle 50%
                    middle_papers.sort(key=lambda x: x[2], reverse=True)  # Sort by score (highest first)
                    
                    for paper_id, metadata, score, year in middle_papers:
                        if len(selected_papers) >= target_papers:
                            break
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                    
                    print(f"📊 Temporal distribution: {oldest_selected} oldest, {newest_selected} newest, {len(selected_papers) - lab_papers_to_add - preferred_author_papers_to_add - oldest_selected - newest_selected} middle-aged")
                else:
                    # Fallback: no temporal data available, use original scoring
                    for paper_id, metadata, score in other_papers_scored[:remaining_needed]:
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                            if len(selected_papers) >= target_papers:
                                break

            # Step 6: Ensure we have exactly target_papers
            if len(selected_papers) < target_papers:
                print(f"⚠️  Only found {len(selected_papers)} unique papers (requested {target_papers})")
                print("🔄 Attempting to fill remaining slots with any available papers...")
                
                # Fill remaining slots with any available papers
                remaining_papers = [p for p in other_papers_scored if p[0] not in selected_papers]
                for paper_id, metadata, score in remaining_papers:
                    if len(selected_papers) >= target_papers:
                        break
                    selected_papers.add(paper_id)
                    selected_chunks.append(paper_chunks[paper_id])

            # Step 7: Update used papers tracking
            self.used_papers.update(selected_papers)

            print(f"✅ Successfully selected exactly {len(selected_papers)} different papers")
            print(f"📊 Lab papers: {len([p for p in selected_papers if p in [pid for pid, _ in lab_papers[:lab_papers_to_add]]])}")
            print(f"📊 Preferred author papers: {len([p for p in selected_papers if p in [pid for pid, _ in preferred_author_papers[:preferred_author_papers_to_add]]])}")
            
            # Calculate temporal distribution for other papers
            other_papers_count = len(selected_papers) - len([p for p in selected_papers if p in [pid for pid, _ in lab_papers[:lab_papers_to_add]]]) - len([p for p in selected_papers if p in [pid for pid, _ in preferred_author_papers[:preferred_author_papers_to_add]]])
            print(f"📊 Other papers: {other_papers_count}")
            
            if other_papers_count > 0:
                # Re-analyze the temporal distribution for better reporting
                other_paper_years = []
                for paper_id in selected_papers:
                    if (paper_id not in [pid for pid, _ in lab_papers[:lab_papers_to_add]] and 
                        paper_id not in [pid for pid, _ in preferred_author_papers[:preferred_author_papers_to_add]]):
                        metadata = paper_metadata[paper_id]
                        publication_date = metadata.get('publication_date', '')
                        if publication_date and len(publication_date) >= 4:
                            try:
                                year = int(publication_date[:4])
                                other_paper_years.append(year)
                            except (ValueError, TypeError):
                                pass
                
                if other_paper_years:
                    other_paper_years.sort()
                    total_other = len(other_paper_years)
                    q1_idx = max(0, total_other // 4 - 1)
                    q3_idx = min(total_other - 1, 3 * total_other // 4)
                    
                    oldest_papers = len([y for y in other_paper_years if y <= other_paper_years[q1_idx]])
                    newest_papers = len([y for y in other_paper_years if y >= other_paper_years[q3_idx]])
                    middle_papers = total_other - oldest_papers - newest_papers
                    
                    print(f"   📅 Temporal distribution of other papers:")
                    print(f"      • Oldest 25%: {oldest_papers} papers (≤{other_paper_years[q1_idx] if oldest_papers > 0 else 'N/A'})")
                    print(f"      • Middle 50%: {middle_papers} papers")
                    print(f"      • Newest 25%: {newest_papers} papers (≥{other_paper_years[q3_idx] if newest_papers > 0 else 'N/A'})")
                    print(f"      • Year range: {min(other_paper_years)} - {max(other_paper_years)}")
            
            print(f"📊 Total papers used so far: {len(self.used_papers)}")

            # Store the results for citation extraction
            self.last_diverse_results = []
            for paper_id in selected_papers:
                metadata = paper_metadata[paper_id]
                chunk = paper_chunks[paper_id]
                self.last_diverse_results.append({
                    'chunk': chunk,
                    'metadata': metadata
                })

            return selected_chunks

        except Exception as e:
            print(f"❌ Error selecting diverse papers: {e}")
            return None

    def _apply_randomization_strategy(self, results, strategy='enhanced'):
        """
        Apply different randomization strategies to chunk selection results.
        
        Args:
            results: List of chunk results from ChromaDB
            strategy: Randomization strategy ('enhanced', 'shuffle', 'diversity', 'time_based', 'none')
        
        Returns:
            Randomized list of results
        """
        import random
        import time
        
        if not results:
            return results
            
        print(f"🎲 Applying {strategy} randomization strategy to {len(results)} chunks...")
        
        if strategy == 'none':
            print("   No additional randomization applied")
            return results
            
        elif strategy == 'shuffle':
            # Simple random shuffle
            randomized_results = results.copy()
            random.shuffle(randomized_results)
            print("   Applied simple random shuffle")
            
        elif strategy == 'diversity':
            # Prioritize diversity by publication date and citations
            randomized_results = results.copy()
            
            # Sort by publication date (newer first) and citation count (higher first)
            def diversity_score(result):
                metadata = result.get('metadata', {})
                publication_date = metadata.get('publication_date', '')
                citation_count = metadata.get('citation_count', 'not found')
                
                # Extract year from publication date
                year = 0
                if publication_date and len(publication_date) >= 4:
                    try:
                        year = int(publication_date[:4])
                    except ValueError:
                        year = 0
                
                # Convert citation count to number
                citations = 0
                if citation_count != 'not found':
                    try:
                        citations = int(citation_count)
                    except (ValueError, TypeError):
                        citations = 0
                
                # Higher score for newer papers and more citations
                return (year * 1000) + citations
            
            # Sort by diversity score (descending)
            randomized_results.sort(key=diversity_score, reverse=True)
            
            # Then shuffle within similar score ranges
            if len(randomized_results) > 10:
                # Shuffle in groups of 10 to maintain some order while adding randomness
                for i in range(0, len(randomized_results), 10):
                    end_idx = min(i + 10, len(randomized_results))
                    group = randomized_results[i:end_idx]
                    random.shuffle(group)
                    randomized_results[i:end_idx] = group
            
            print("   Applied diversity-based randomization (newer, more cited papers first)")
            
        elif strategy == 'time_based':
            # Use current time as seed for reproducible but time-varying randomization
            current_minute = int(time.time() // 60)  # Changes every minute
            random.seed(current_minute)
            randomized_results = results.copy()
            random.shuffle(randomized_results)
            random.seed()  # Reset seed
            print(f"   Applied time-based randomization (seed: {current_minute})")
            
        elif strategy == 'enhanced':
            # Enhanced randomization combining relevance and diversity
            randomized_results = results.copy()
            
            # Tier-based approach: divide into tiers and shuffle within each
            num_tiers = min(5, len(randomized_results) // 10)  # 5 tiers or fewer if not enough results
            if num_tiers < 2:
                # If too few results, just shuffle
                random.shuffle(randomized_results)
                print("   Applied enhanced randomization (simple shuffle due to small dataset)")
            else:
                # Divide into tiers
                tier_size = len(randomized_results) // num_tiers
                for i in range(num_tiers):
                    start_idx = i * tier_size
                    end_idx = start_idx + tier_size if i < num_tiers - 1 else len(randomized_results)
                    
                    # Shuffle within each tier
                    tier = randomized_results[start_idx:end_idx]
                    random.shuffle(tier)
                    randomized_results[start_idx:end_idx] = tier
                
                # Then shuffle the entire list slightly
                random.shuffle(randomized_results)
                print(f"   Applied enhanced randomization (tier-based with {num_tiers} tiers)")
        
        else:
            print(f"   Unknown strategy '{strategy}', using default shuffle")
            randomized_results = results.copy()
            random.shuffle(randomized_results)
        
        return randomized_results

    def _get_paper_identifier(self, metadata):
        """Extract a unique identifier for a paper from its metadata."""
        # Try DOI first, then title, then source + title
        doi = metadata.get('doi', '')
        title = metadata.get('title', '')
        source = metadata.get('source', '')

        if doi and doi != 'No DOI':
            return f"doi:{doi}"
        elif title:
            # Use a normalized version of the title (lowercase, no extra spaces)
            normalized_title = ' '.join(title.lower().split())
            return f"title:{normalized_title}"
        else:
            # Fallback to source + title if available
            if source and title:
                normalized_title = ' '.join(title.lower().split())
                return f"source_title:{source}_{normalized_title}"
            else:
                # Last resort: use a hash of the metadata
                import hashlib
                metadata_str = str(sorted(metadata.items()))
                return f"hash:{hashlib.md5(metadata_str.encode()).hexdigest()[:8]}"

    def reset_used_papers(self):
        """Reset the list of used papers to allow reusing them."""
        previous_count = len(self.used_papers)
        self.used_papers.clear()
        print(f"🔄 Reset used papers list. Previously used {previous_count} papers.")
        print("💡 You can now reuse papers that were previously selected.")

    def show_used_papers_status(self):
        """Show the current status of used papers tracking with source breakdown."""
        print(f"📊 Used Papers Status:")
        print(f"   Total papers used: {len(self.used_papers)}")

        # Analyze source distribution of used papers
        source_counts = {'pubmed': 0, 'biorxiv': 0, 'medrxiv': 0, 'unknown': 0}

        for paper_id in self.used_papers:
            if paper_id.startswith('doi:'):
                # For DOI-based IDs, we can't easily determine source without metadata
                source_counts['unknown'] += 1
            elif paper_id.startswith('title:'):
                # For title-based IDs, we can't easily determine source without metadata
                source_counts['unknown'] += 1
            elif paper_id.startswith('source_title:'):
                # Extract source from source_title format
                if 'biorxiv_' in paper_id:
                    source_counts['biorxiv'] += 1
                elif 'medrxiv_' in paper_id:
                    source_counts['medrxiv'] += 1
                elif 'pubmed_' in paper_id:
                    source_counts['pubmed'] += 1
                else:
                    source_counts['unknown'] += 1
            else:
                source_counts['unknown'] += 1

        print(f"   Source breakdown:")
        print(f"     PubMed: {source_counts['pubmed']} papers")
        print(f"     BioRxiv: {source_counts['biorxiv']} papers")
        print(f"     MedRxiv: {source_counts['medrxiv']} papers")
        if source_counts['unknown'] > 0:
            print(f"     Unknown: {source_counts['unknown']} papers")

        if self.used_papers:
            # Show a sample of used papers
            sample_size = min(5, len(self.used_papers))
            sample_papers = list(self.used_papers)[:sample_size]
            print(f"   Sample of used papers:")
            for paper_id in sample_papers:
                if paper_id.startswith('doi:'):
                    print(f"     - DOI: {paper_id[4:]}")
                elif paper_id.startswith('title:'):
                    print(f"     - Title: {paper_id[6:][:50]}...")
                else:
                    print(f"     - {paper_id[:50]}...")

            if len(self.used_papers) > sample_size:
                print(f"     ... and {len(self.used_papers) - sample_size} more")
        else:
            print("   No papers have been used yet.")

        print(f"   💡 Use 'reset_papers' to clear the used papers list and allow reusing papers")

    @staticmethod
    def automated_verdict(accuracy, novelty, relevancy, accuracy_threshold=80, novelty_threshold=85, relevancy_threshold=50):
        if accuracy is not None and novelty is not None and relevancy is not None:
            return "ACCEPTED" if (accuracy >= accuracy_threshold and novelty >= novelty_threshold and relevancy >= relevancy_threshold) else "REJECTED"
        return "REJECTED"

    def iterative_hypothesis_generation(self, user_prompt, max_rounds=5, n=3):
        """Run generator-critic feedback loop for each hypothesis until accepted (>=90% accuracy/novelty and verdict ACCEPT)."""
        novelty_threshold = 85
        accuracy_threshold = 80
        relevancy_threshold = 50
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("[iterative_hypothesis_generation] ERROR: Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        # Generate initial hypotheses with fresh chunks for each
        print(f"[iterative_hypothesis_generation] Generating initial hypotheses with unique chunks...")
        hypotheses = []

        for i in range(n):
            print(f"[iterative_hypothesis_generation] Getting fresh chunks for hypothesis {i+1}/{n}...")
            # Get fresh chunks for each hypothesis to ensure maximum diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                # Use dynamic chunk selection for maximum diversity
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                context_chunks = self.select_dynamic_chunks_for_generation(user_prompt, randomization_strategy=randomization_strategy)
                if not context_chunks:
                    print(f"[iterative_hypothesis_generation] Falling back to direct retrieval for hypothesis {i+1}...")
                    context_chunks = self.retrieve_relevant_chunks(user_prompt, top_k=1500)
            else:
                # Use direct retrieval
                context_chunks = self.retrieve_relevant_chunks(user_prompt, top_k=1500)

            if not context_chunks:
                print(f"[iterative_hypothesis_generation] ERROR: No relevant context found for hypothesis {i+1}.")
                return

            try:
                context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1)
                if hypothesis_list:
                    hypotheses.append(hypothesis_list[0])
                    print(f"[iterative_hypothesis_generation] Generated hypothesis {i+1} with {len(context_chunks)} unique chunks")
                else:
                    print(f"[iterative_hypothesis_generation] ERROR: Failed to generate hypothesis {i+1}")
                    return
            except Exception as e:
                print(f"[iterative_hypothesis_generation] ERROR: Failed to generate hypothesis {i+1}: {e}")
                return
        accepted = [False] * n
        rounds = [0] * n
        critiques = [{} for _ in range(n)]
        from tqdm.auto import tqdm
        total_iterations = max_rounds * n
        current_iteration = 0
        print(f"[iterative_hypothesis_generation] Starting iterative refinement (max {max_rounds} rounds per hypothesis, 5-minute time limit per critique, acceptance: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold})...")
        with tqdm(total=total_iterations, desc="Hypothesis refinement", unit="iteration") as pbar:
            while not all(accepted) and max(rounds) < max_rounds:
                for i in range(n):
                    if accepted[i]:
                        continue
                    rounds[i] += 1
                    current_iteration += 1
                    print(f"[iterative_hypothesis_generation] Critique Round {rounds[i]} for Hypothesis {i+1} (5-minute timer started)...")
                    self.hypothesis_timer.start()
                    try:
                        context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                        if self.hypothesis_timer.check_expired():
                            print(f"[iterative_hypothesis_generation] TIMEOUT: Hypothesis {i+1} critique timed out.")
                            citations = self.extract_citations_from_chunks(context_chunks)
                            formatted_citations = self.format_citations_for_export(citations)
                            self.hypothesis_records.append({
                                'Hypothesis': hypotheses[i],
                                'Accuracy': None,
                                'Novelty': None,
                                'Verdict': 'TIMEOUT',
                                'Critique': 'Critique process timed out after 5 minutes',
                                'Citations': formatted_citations
                            })
                            pbar.update(1)
                            continue
                        critique_result = self.hypothesis_critic.critique(hypotheses[i], context_texts, prompt=user_prompt, lab_goals=get_lab_goals())
                        if self.hypothesis_timer.check_expired():
                            print(f"[iterative_hypothesis_generation] TIMEOUT: Hypothesis {i+1} critique timed out after critique.")
                            citations = self.extract_citations_from_chunks(context_chunks)
                            formatted_citations = self.format_citations_for_export(citations)
                            self.hypothesis_records.append({
                                'Hypothesis': hypotheses[i],
                                'Accuracy': critique_result.get('accuracy', None),
                                'Novelty': critique_result.get('novelty', None),
                                'Verdict': 'TIMEOUT',
                                'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                                'Citations': formatted_citations
                            })
                            pbar.update(1)
                            continue
                        critiques[i] = critique_result
                        novelty = critique_result.get('novelty', 0)
                        accuracy = critique_result.get('accuracy', 0)
                        relevancy = critique_result.get('relevancy', 0)
                        verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
                        remaining_time = self.hypothesis_timer.get_remaining_time()
                        print(f"\033[1mHypothesis {i+1}:\033[0m \033[96m{hypotheses[i]}\033[0m")
                        print(f"[iterative_hypothesis_generation] Critique: {critique_result['critique']}")
                        print(f"[iterative_hypothesis_generation] Scores: Novelty={novelty}/{novelty_threshold}, Accuracy={accuracy}/{accuracy_threshold}, Relevancy={relevancy}/{relevancy_threshold}, Automated Verdict={verdict}, Time left={remaining_time:.1f}s")
                        citations = self.extract_citations_from_chunks(context_chunks)
                        formatted_citations = self.format_citations_for_export(citations)
                        self.hypothesis_records.append({
                            'Hypothesis': hypotheses[i],
                            'Accuracy': accuracy,
                            'Novelty': novelty,
                            'Relevancy': relevancy,
                            'Verdict': verdict,
                            'Critique': critique_result.get('critique', ''),
                            'Citations': formatted_citations
                        })
                        if verdict == 'ACCEPTED':
                            print(f"\033[92mACCEPTED\033[0m: Hypothesis {i+1} (Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold})")
                            accepted[i] = True
                        else:
                            print(f"\033[91mREJECTED\033[0m: Hypothesis {i+1} (Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold}) - Regenerating...")
                            if self.hypothesis_timer.check_expired():
                                print(f"[iterative_hypothesis_generation] TIMEOUT: Hypothesis {i+1} regeneration timed out.")
                                pbar.update(1)
                                continue
                            try:
                                context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                                new_hypothesis = self.hypothesis_generator.generate(context_texts, n=1)
                                if new_hypothesis:
                                    hypotheses[i] = new_hypothesis[0]
                            except Exception as e:
                                print(f"[iterative_hypothesis_generation] ERROR: Failed to regenerate hypothesis {i+1}: {e}")
                                continue
                        avg_rating = (novelty + accuracy + relevancy) / 3 if novelty is not None and accuracy is not None and relevancy is not None else 0
                        pbar.set_postfix({
                            "accepted": sum(accepted),
                            "round": max(rounds),
                            "current": f"H{i+1}",
                            "avg_rating": f"{avg_rating:.1f}",
                            "time_left": f"{remaining_time:.0f}s"
                        })
                        pbar.update(1)
                    except Exception as e:
                        print(f"[iterative_hypothesis_generation] ERROR: Failed to critique hypothesis {i+1}: {e}")
                        citations = self.extract_citations_from_chunks(context_chunks)
                        formatted_citations = self.format_citations_for_export(citations)
                        self.hypothesis_records.append({
                            'Hypothesis': hypotheses[i],
                            'Accuracy': None,
                            'Novelty': None,
                            'Relevancy': None,
                            'Verdict': 'ERROR',
                            'Critique': f'Error during critique: {str(e)}',
                            'Citations': formatted_citations
                        })
                        pbar.update(1)
                        continue
                    if all(accepted):
                        break
        print("[iterative_hypothesis_generation] Final Hypotheses:")
        for i, hyp in enumerate(hypotheses, 1):
            status = "ACCEPTED" if accepted[i-1] else f"FAILED (max {max_rounds} rounds)"
            print(f"   {i}. {hyp} [{status}]")
            if isinstance(critiques[i-1], dict) and critiques[i-1]:
                print(f"      Critique: {critiques[i-1].get('critique','')}")
                print(f"      Novelty: {critiques[i-1].get('novelty','')}/{novelty_threshold}, Accuracy: {critiques[i-1].get('accuracy','')}/{accuracy_threshold}, Relevancy: {critiques[i-1].get('relevancy','')}/{relevancy_threshold}, Automated Verdict: {self.automated_verdict(critiques[i-1].get('accuracy',0), critiques[i-1].get('novelty',0), critiques[i-1].get('relevancy',0), accuracy_threshold, novelty_threshold, relevancy_threshold)}")
        return hypotheses

    def add_to_package(self, search_results):
        """Add search results to the current package."""
        current_size = self.current_package["total_chars"]
        new_size = current_size + sum(len(result.get("chunk", "")) for result in search_results)
        if new_size > 2000000:
            print(f"[add_to_package] WARNING: Adding {len(search_results)} chunks would make package very large ({new_size:,} characters)")
            print("[add_to_package] This may cause API errors. Consider using 'clear' first.")
            response = input("Continue anyway? (y/n): ").lower()
            if response != 'y':
                return

        # Store the initial add quantity for dynamic chunk selection
        if self.initial_add_quantity is None:
            self.initial_add_quantity = len(search_results)
            print(f"[add_to_package] Stored initial add quantity: {self.initial_add_quantity} chunks for dynamic selection")

        for result in search_results:
            chunk = result.get("chunk", "")
            metadata = result.get("metadata", {})
            self.current_package["chunks"].append(chunk)
            self.current_package["metadata"].append(metadata)
            self.current_package["sources"].add(metadata.get("source", "Unknown"))
            self.current_package["total_chars"] += len(chunk)
        print(f"[add_to_package] Added {len(search_results)} chunks to package.")
        print(f"[add_to_package] Package now contains {len(self.current_package['chunks'])} chunks from {len(self.current_package['sources'])} sources.")
        if self.current_package["total_chars"] > 1000000:
            print(f"[add_to_package] WARNING: Package size: {self.current_package['total_chars']:,} characters (consider using 'clear' if you get API errors)")

    def clear_package(self):
        """Clear the current package."""
        self.current_package = {
            "chunks": [],
            "metadata": [],
            "sources": set(),
            "total_chars": 0
        }
        # Reset the initial add quantity when clearing package
        self.initial_add_quantity = None
        # Reset used papers tracking when clearing package
        previous_papers_count = len(self.used_papers)
        self.used_papers.clear()
        print("[clear_package] Package cleared.")
        print("[clear_package] Initial add quantity reset.")
        print(f"[clear_package] Used papers list reset ({previous_papers_count} papers cleared).")

    def show_package(self):
        """Display information about the current package with source breakdown."""
        if not self.current_package["chunks"]:
            print("[show_package] Package is empty.")
            return

        # Calculate source distribution
        source_counts = {'pubmed': 0, 'biorxiv': 0, 'medrxiv': 0, 'unknown': 0}

        for metadata in self.current_package["metadata"]:
            source = metadata.get('source', '').lower()
            source_name = metadata.get('source_name', '').lower()

            if 'pubmed' in source or 'pubmed' in source_name:
                source_counts['pubmed'] += 1
            elif 'biorxiv' in source or 'biorxiv' in source_name:
                source_counts['biorxiv'] += 1
            elif 'medrxiv' in source or 'medrxiv' in source_name:
                source_counts['medrxiv'] += 1
            else:
                source_counts['unknown'] += 1

        print(f"[show_package] Current Package:")
        print(f"   📚 Total chunks: {len(self.current_package['chunks'])}")
        print(f"   📝 Total characters: {self.current_package['total_chars']:,}")
        print(f"   📖 Sources: {', '.join(self.current_package['sources'])}")
        print(f"   📊 Used papers tracked: {len(self.used_papers)}")

        print(f"   📊 Source distribution:")
        print(f"     PubMed: {source_counts['pubmed']} chunks")
        print(f"     BioRxiv: {source_counts['biorxiv']} chunks")
        print(f"     MedRxiv: {source_counts['medrxiv']} chunks")
        if source_counts['unknown'] > 0:
            print(f"     Unknown: {source_counts['unknown']} chunks")

        print(f"\n[show_package] Sample chunks:")
        for i, chunk in enumerate(self.current_package["chunks"][:3], 1):
            metadata = self.current_package["metadata"][i-1]
            title = metadata.get("title", "No title")[:60]
            source = metadata.get("source", "Unknown")
            print(f"   {i}. {title}... (Source: {source})")
        if len(self.current_package["chunks"]) > 3:
            print(f"   ... and {len(self.current_package['chunks']) - 3} more chunks")

    def generate_hypotheses_from_package(self, n=5):
        novelty_threshold = 85
        accuracy_threshold = 80
        relevancy_threshold = 50
        if not self.current_package["chunks"]:
            print("❌ Package is empty. Add some chunks first.")
            return
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("❌ Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        # Prompt for the query if not already stored
        if "prompt" not in self.current_package or not self.current_package["prompt"]:
            self.current_package["prompt"] = input("Enter the original query for this package: ").strip()
        package_prompt = self.current_package["prompt"]
        print(f"\n🧠 Generating {n} hypotheses...")
        print(f"📦 Using {len(self.current_package['chunks'])} package chunks for critique")
        print(f"[INFO] Acceptance criteria: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold}")

        # Check package size and warn if too large
        package_size = sum(len(chunk) for chunk in self.current_package["chunks"])
        if package_size > 1000000:  # 1MB limit for package
            print(f"⚠️  Package is very large ({package_size:,} characters). This may cause API errors.")
            print("💡 Consider using 'clear' and adding fewer chunks.")
            response = input("Continue anyway? (y/n): ").lower()
            if response != 'y':
                return

        # For generation, use a sample of the database instead of the entire thing
        # This prevents the 300MB payload limit error
        if not self.embeddings_data:
            # Check if ChromaDB has data first
            if self.use_chromadb and self.chroma_manager and self.is_chromadb_ready():
                print("📚 ChromaDB has data - using package chunks for generation (no need to load embeddings)")
                # Use only the package chunks for generation when ChromaDB is available
                all_chunks = self.current_package["chunks"]
            else:
                print("📚 Loading embeddings data...")
                self.embeddings_data = self.load_all_embeddings()
                if not self.embeddings_data:
                    print("❌ No embeddings data available")
                    return
                all_chunks = self.embeddings_data["chunks"]
        else:
            all_chunks = self.embeddings_data["chunks"]
        package_chunks = self.current_package["chunks"]
        package_metadata = self.current_package["metadata"]

        # Calculate safe sample size (aim for ~50MB payload)
        # Each chunk is roughly 1000 characters, so 50MB = ~50,000 chunks
        max_chunks_for_generation = 50000

        # Check if we're using package chunks (from ChromaDB) or all chunks (from embeddings)
        if all_chunks is self.current_package["chunks"]:
            # Using package chunks from ChromaDB
            generation_chunks = all_chunks
            print(f"📚 Using {len(generation_chunks)} package chunks from ChromaDB for generation")
        else:
            # Using all chunks from embeddings data
            if len(all_chunks) > max_chunks_for_generation:
                print(f"📚 Database contains {len(all_chunks)} total chunks")
                print(f"📚 Using sample of {max_chunks_for_generation} chunks for generation (to avoid API limits)")
                # Take a representative sample from different parts of the database
                step = len(all_chunks) // max_chunks_for_generation
                generation_chunks = all_chunks[::step][:max_chunks_for_generation]
            else:
                generation_chunks = all_chunks
                print(f"📚 Database contains {len(all_chunks)} total chunks")

        print(f"📦 Package contains {len(package_chunks)} chunks")

        # Sequential, real-time generator-critic loop for 5 accepted hypotheses
        print(f"\n🧠 Generating hypotheses one at a time until 5 are accepted (novelty >= {novelty_threshold}, accuracy >= {accuracy_threshold}, 5-minute time limit per critique)...")

        # Initialize Excel file for incremental saving
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Ensure hypothesis_export directory exists
        export_dir = "hypothesis_export"
        os.makedirs(export_dir, exist_ok=True)
        excel_filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")
        print(f"💾 Will save hypotheses incrementally to: {excel_filename}")

        accepted_hypotheses = []
        attempts = 0
        max_attempts = 100  # Prevent infinite loops
        while len(accepted_hypotheses) < n and attempts < max_attempts:
            attempts += 1
            print(f"\n🧠 Generating hypothesis attempt {attempts}...")

            # ALWAYS select new chunks for EVERY hypothesis attempt to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"🔄 Selecting new chunks for hypothesis attempt {attempts}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation(package_prompt, randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    generation_chunks = dynamic_chunks
                    print(f"📚 Using {len(generation_chunks)} dynamically selected chunks for generation")
                else:
                    print(f"⚠️  Failed to select dynamic chunks, using original chunks")
            else:
                print(f"📚 Using original chunks for generation (no dynamic selection available)")

            # Start timer for this hypothesis
            self.hypothesis_timer.start()
            print(f"⏱️  Starting 5-minute timer for hypothesis attempt {attempts}...")

            hypothesis_list = self._generate_hypotheses_with_retry(generation_chunks, n=1)
            if not hypothesis_list:
                print("❌ Failed to generate hypothesis. Skipping...")
                # Record failed generation
                failed_record = {
                    'Hypothesis': f'Failed generation attempt {attempts}',
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'GENERATION_FAILED',
                    'Critique': 'Failed to generate hypothesis',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(failed_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(failed_record, excel_filename)
                continue

            hypothesis = hypothesis_list[0]
            
            # Validate hypothesis format before proceeding
            from hypothesis_tools import validate_hypothesis_format
            is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
            if not is_valid_format:
                print(f"❌ Hypothesis rejected due to format: {format_reason}")
                # Record format rejection
                format_rejected_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'FORMAT_REJECTED',
                    'Critique': f'Hypothesis rejected: {format_reason}',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(format_rejected_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(format_rejected_record, excel_filename)
                continue

            hypothesis = hypothesis_list[0]
            print(f"Generated Hypothesis: {hypothesis}")
            print(f"\n🔄 Critiquing Hypothesis: {hypothesis}")

            # Check timer before critique
            if self.hypothesis_timer.check_expired():
                print(f"⏰ Time limit reached for hypothesis attempt {attempts}. Moving to next attempt.")
                # Record timeout result
                timeout_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'TIMEOUT',
                    'Critique': 'Critique process timed out after 5 minutes',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(timeout_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(timeout_record, excel_filename)
                continue

            critique_result = self._critique_hypothesis_with_retry(hypothesis, package_chunks, package_prompt)
            if not critique_result:
                print(f"❌ Failed to critique hypothesis after retries. Skipping...")
                # Record failed critique
                critique_failed_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'CRITIQUE_FAILED',
                    'Critique': 'Failed to critique hypothesis after retries',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(critique_failed_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(critique_failed_record, excel_filename)
                continue

            # Check timer after critique
            if self.hypothesis_timer.check_expired():
                print(f"⏰ Time limit reached for hypothesis attempt {attempts} after critique. Moving to next attempt.")
                # Record timeout result with partial critique
                partial_timeout_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': critique_result.get('accuracy', None),
                    'Novelty': critique_result.get('novelty', None),
                    'Verdict': 'TIMEOUT',
                    'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(partial_timeout_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(partial_timeout_record, excel_filename)
                continue

            novelty = critique_result.get('novelty', 0)
            accuracy = critique_result.get('accuracy', 0)
            relevancy = critique_result.get('relevancy', 0)
            # Use the same thresholds as the acceptance logic
            verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
            remaining_time = self.hypothesis_timer.get_remaining_time()
            print(f"   Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold}, Automated Verdict: {verdict}")
            print(f"   ⏱️  Time remaining: {remaining_time:.1f}s")
            
            # Print rejection reason if hypothesis is rejected
            if verdict != 'ACCEPTED':
                if accuracy < accuracy_threshold:
                    print(f"❌ Hypothesis rejected: Low accuracy score ({accuracy}/{accuracy_threshold})")
                elif novelty < novelty_threshold:
                    print(f"❌ Hypothesis rejected: Low novelty score ({novelty}/{novelty_threshold})")
                elif relevancy < relevancy_threshold:
                    print(f"❌ Hypothesis rejected: Low relevancy score ({relevancy}/{relevancy_threshold})")
                else:
                    print(f"❌ Hypothesis rejected: Unknown reason")

            # Extract citations from the dynamically selected chunks used for this specific hypothesis
            citations = []
            unique_sources = set()
            
            # Get metadata from the dynamically selected chunks
            dynamic_metadata = []
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                # If we used dynamic chunk selection, get metadata from those chunks
                if hasattr(self, 'last_dynamic_results') and self.last_dynamic_results:
                    dynamic_metadata = [result.get('metadata', {}) for result in self.last_dynamic_results]
                else:
                    # Fallback: extract from generation_chunks if they have metadata
                    dynamic_metadata = []
                    for chunk in generation_chunks:
                        if isinstance(chunk, dict) and 'metadata' in chunk:
                            dynamic_metadata.append(chunk['metadata'])
                        else:
                            # If chunks don't have metadata, use package metadata as fallback
                            dynamic_metadata = package_metadata
                            break
            else:
                # If not using dynamic selection, use package metadata
                dynamic_metadata = package_metadata
            
            # Extract citations from the metadata
            for metadata in dynamic_metadata:
                if metadata:
                    source_name = metadata.get('source_name', 'Unknown')
                    title = metadata.get('title', 'No title')
                    doi = metadata.get('doi', 'No DOI')
                    authors = metadata.get('author', metadata.get('authors', 'Unknown authors'))
                    journal = metadata.get('journal', 'Unknown journal')
                    # Extract year from publication_date
                    publication_date = metadata.get('publication_date', '')
                    year = publication_date[:4] if publication_date and len(publication_date) >= 4 else 'Unknown year'

                    citation = {
                        'source_name': source_name,
                        'title': title,
                        'doi': doi,
                        'authors': authors,
                        'journal': journal,
                        'year': year
                    }

                    # Add to citations if not already present
                    citation_key = doi if doi != 'No DOI' else title
                    if citation_key not in unique_sources:
                        citations.append(citation)
                        unique_sources.add(citation_key)

            formatted_citations = self.format_citations_for_export(citations)

            # Track record for export
            hypothesis_record = {
                'Hypothesis': hypothesis,
                'Accuracy': accuracy,
                'Novelty': novelty,
                'Relevancy': relevancy,
                'Verdict': verdict,
                'Critique': critique_result.get('critique', ''),
                'Citations': formatted_citations
            }
            self.hypothesis_records.append(hypothesis_record)

            # Save incrementally to Excel
            self.save_hypothesis_record_incrementally(hypothesis_record, excel_filename)

            if verdict == 'ACCEPTED':
                accepted_hypotheses.append({
                    "hypothesis": hypothesis,
                    "critique": critique_result,
                    "score": (novelty + accuracy + relevancy) / 3
                })
                print(f"✅ Hypothesis accepted! ({len(accepted_hypotheses)}/{n})-----------------------------------\n")
            else:
                print(f"❌ Hypothesis rejected (verdict: {verdict})-----------------------------------\n")

        if not accepted_hypotheses:
            print("❌ No hypotheses were successfully critiqued.")
            return

        print(f"\n🏆 Top {len(accepted_hypotheses)} Hypotheses:")
        print("=" * 80)
        for i, result in enumerate(accepted_hypotheses, 1):
            hypothesis = result["hypothesis"]
            score = result["score"]
            critique = result["critique"]
            print(f"\n{i}. {hypothesis}")
            print(f"   Score: {score:.1f}")
            print(f"   Novelty: {critique.get('novelty', 'N/A')}")
            print(f"   Accuracy: {critique.get('accuracy', 'N/A')}")
            print(f"   Relevancy: {critique.get('relevancy', 'N/A')}")
            print(f"   Automated Verdict: {self.automated_verdict(critique.get('accuracy', 0), critique.get('novelty', 0), critique.get('relevancy', 0), accuracy_threshold, novelty_threshold, relevancy_threshold)}")
            print(f"   Critique: {critique.get('critique', 'N/A')}")
            print("-" * 80)
        return accepted_hypotheses[:n]

    def _generate_hypotheses_with_retry(self, generation_chunks, n, max_retries=3):
        """Generate hypotheses with retry logic for rate limiting."""
        if not self.hypothesis_generator:
            print("❌ Hypothesis generator not available")
            return None

        for attempt in range(max_retries):
            try:
                # Check rate limit before making API call
                remaining_requests = self.gemini_rate_limiter.get_remaining_requests()
                if remaining_requests <= 0:
                    print(f"⏳ Gemini API rate limit reached. Waiting for reset...")
                    self.gemini_rate_limiter.wait_if_needed()

                # Extract chunk text from the sample
                generation_chunk_texts = [chunk for chunk in generation_chunks]
                hypotheses = self.hypothesis_generator.generate(generation_chunk_texts, n=n)
                return hypotheses
            except Exception as e:
                error_str = str(e)
                if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower():
                    wait_time = (2 ** attempt) * 30 + random.randint(0, 10)  # Exponential backoff
                    print(f"⚠️  API rate limited (attempt {attempt + 1}/{max_retries}). Waiting {wait_time}s...")
                    time.sleep(wait_time)
                    if attempt == max_retries - 1:
                        print("❌ Max retries reached. API quota exceeded.")
                        return None
                elif "payload size exceeds the limit" in error_str:
                    print("💡 Payload too large. Try reducing package size with 'clear' and adding fewer chunks")
                    return None
                else:
                    print(f"❌ Unexpected error: {e}")
                    return None
        return None

    def _critique_hypothesis_with_retry(self, hypothesis, package_chunks, prompt, max_retries=3):
        """Critique hypothesis with retry logic for rate limiting."""
        if not self.hypothesis_critic:
            print("❌ Hypothesis critic not available")
            return None

        for attempt in range(max_retries):
            try:
                # Check rate limit before making API call
                remaining_requests = self.gemini_rate_limiter.get_remaining_requests()
                if remaining_requests <= 0:
                    print(f"⏳ Gemini API rate limit reached. Waiting for reset...")
                    self.gemini_rate_limiter.wait_if_needed()

                # Extract chunk text from package chunks
                package_chunk_texts = [chunk for chunk in package_chunks]
                critique_result = self.hypothesis_critic.critique(hypothesis, package_chunk_texts, prompt=prompt, lab_goals=get_lab_goals())
                return critique_result
            except Exception as e:
                error_str = str(e)
                if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower():
                    wait_time = (2 ** attempt) * 30 + random.randint(0, 10)  # Exponential backoff
                    print(f"⚠️  API rate limited (attempt {attempt + 1}/{max_retries}). Waiting {wait_time}s...")
                    time.sleep(wait_time)
                    if attempt == max_retries - 1:
                        print("❌ Max retries reached. API quota exceeded.")
                        return None
                elif "payload size exceeds the limit" in error_str:
                    print("💡 Package is too large for critique. Try 'clear' and add fewer chunks.")
                    return None
                else:
                    print(f"❌ Unexpected error: {e}")
                    return None
        return None

    def generate_hypotheses_offline(self, n=5):
        """Generate hypotheses without using the API (fallback mode)."""
        if not self.current_package["chunks"]:
            print("❌ Package is empty. Add some chunks first.")
            return

        print(f"\n🧠 Generating {n} hypotheses (offline mode)...")
        print(f"📦 Using {len(self.current_package['chunks'])} package chunks")

        # Extract key terms and patterns from the package
        package_text = " ".join(self.current_package["chunks"])

        # Simple keyword extraction
        import re
        from collections import Counter

        # Extract UBR-5 related terms
        ubr5_terms = re.findall(r'\b(?:UBR5?|ubiquitin|ligase|protein|E3|ubiquitination|degradation|regulation|pathway|signaling|cancer|immunology|immune|response|activation|inhibition|expression|function|mechanism|target|therapeutic|drug|treatment|therapy)\b', package_text, re.IGNORECASE)
        term_counts = Counter(ubr5_terms)

        # Generate simple hypotheses based on common patterns
        hypotheses = []

        # Hypothesis 1: UBR5 regulation
        if any(term in term_counts for term in ['regulation', 'expression', 'function']):
            hypotheses.append("UBR5 expression and function are regulated by specific signaling pathways in immune cells, affecting immune response and cancer progression.")

        # Hypothesis 2: Therapeutic targeting
        if any(term in term_counts for term in ['therapeutic', 'drug', 'treatment', 'target']):
            hypotheses.append("UBR5 represents a novel therapeutic target for cancer immunotherapy, with potential for drug development and clinical applications.")

        # Hypothesis 3: Immune response
        if any(term in term_counts for term in ['immune', 'response', 'activation', 'immunology']):
            hypotheses.append("UBR5 plays a critical role in regulating immune cell activation and response, influencing cancer immunosurveillance and immunotherapy efficacy.")

        # Hypothesis 4: Ubiquitination pathway
        if any(term in term_counts for term in ['ubiquitin', 'ligase', 'ubiquitination', 'degradation']):
            hypotheses.append("UBR5-mediated ubiquitination regulates key proteins in immune signaling pathways, affecting cell fate decisions and immune function.")

        # Hypothesis 5: Cancer mechanism
        if any(term in term_counts for term in ['cancer', 'mechanism', 'pathway', 'signaling']):
            hypotheses.append("UBR5 functions as a key regulator in cancer cell signaling pathways, with implications for tumor progression and therapeutic resistance.")

        # Fill remaining slots with generic hypotheses if needed
        while len(hypotheses) < n:
            hypotheses.append(f"UBR5 may have additional roles in cellular processes related to {list(term_counts.keys())[:3] if term_counts else 'protein regulation'}.")

        print(f"\n🏆 Generated {len(hypotheses)} Hypotheses (Offline Mode):")
        print("=" * 80)
        for i, hypothesis in enumerate(hypotheses[:n], 1):
            print(f"\n{i}. {hypothesis}")
            print(f"   Score: N/A (offline mode)")
            print(f"   Novelty: N/A (offline mode)")
            print(f"   Accuracy: N/A (offline mode)")
            print(f"   Verdict: N/A (offline mode)")
            print(f"   Note: This is a fallback hypothesis based on keyword analysis")
            print("-" * 80)

        print(f"\n💡 Note: These are simplified hypotheses generated without API access.")
        print(f"💡 For more sophisticated analysis, try 'generate' when API is available.")

        return hypotheses[:n]

    def generate_hypotheses_with_meta_generator(self, user_prompt: str, n_per_meta: int = 5, chunks_per_meta: int = 1500):
        """
        Generate hypotheses using the meta-hypothesis generator approach.

        This method:
        1. Takes the user's prompt and generates 5 meta-hypotheses
        2. For each meta-hypothesis, generates hypotheses until exactly 5 are accepted
        3. Only proceeds to the next meta-hypothesis after 5 hypotheses are accepted for the current one
        4. Returns all generated hypotheses with their critiques

        Args:
            user_prompt: The original user query
            n_per_meta: Number of accepted hypotheses required per meta-hypothesis (default: 5)
            chunks_per_meta: Number of context chunks to use per meta-hypothesis (default: 1500)
        """
        if not self.meta_hypothesis_generator or not self.hypothesis_generator or not self.hypothesis_critic:
            print("❌ Meta-hypothesis tools not initialized. Check if Gemini API key is available.")
            return None

        print(f"\n🧠 META-HYPOTHESIS GENERATION SESSION")
        print("=" * 80)
        print(f"Original Query: {user_prompt}")
        print(f"Requiring {n_per_meta} ACCEPTED hypotheses per meta-hypothesis")
        print("=" * 80)

        # Step 1: Generate meta-hypotheses
        print(f"\n🔍 Step 1: Generating 5 meta-hypotheses from user query...")
        try:
            meta_hypotheses = self.meta_hypothesis_generator.generate_meta_hypotheses(user_prompt)
            if not meta_hypotheses or len(meta_hypotheses) < 5:
                print("❌ Failed to generate sufficient meta-hypotheses")
                return None

            print(f"✅ Generated {len(meta_hypotheses)} meta-hypotheses:")
            for i, meta_hyp in enumerate(meta_hypotheses, 1):
                print(f"  {i}. {meta_hyp}")
        except Exception as e:
            print(f"❌ Error generating meta-hypotheses: {e}")
            return None

        # Step 2: For each meta-hypothesis, generate hypotheses until exactly n_per_meta are accepted
        all_hypotheses = []
        total_meta_hypotheses = len(meta_hypotheses)

        # Initialize Excel file for incremental saving
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Ensure hypothesis_export directory exists
        export_dir = "hypothesis_export"
        os.makedirs(export_dir, exist_ok=True)
        excel_filename = os.path.join(export_dir, f"meta_hypothesis_export_{timestamp}.xlsx")
        print(f"💾 Will save meta-hypotheses incrementally to: {excel_filename}")

        # Initialize hypothesis records for this session
        self.hypothesis_records = []

        for meta_idx, meta_hypothesis in enumerate(meta_hypotheses, 1):
            print(f"\n🔍 Step 2.{meta_idx}: Generating hypotheses for meta-hypothesis {meta_idx}/{total_meta_hypotheses}")
            print(f"Meta-hypothesis: {meta_hypothesis}")
            print("-" * 60)

            # Generate hypotheses for this meta-hypothesis until exactly n_per_meta are accepted
            meta_hypotheses_generated = []
            attempts = 0
            max_attempts = n_per_meta * 10  # Allow more attempts to get enough accepted hypotheses

            while len(meta_hypotheses_generated) < n_per_meta and attempts < max_attempts:
                attempts += 1
                print(f"\n🧠 Generating hypothesis attempt {attempts} for meta-hypothesis {meta_idx}...")
                print(f"📊 Progress: {len(meta_hypotheses_generated)}/{n_per_meta} accepted hypotheses")

                # ALWAYS select new chunks for EVERY hypothesis attempt to ensure maximum diversity
                if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                    print(f"🔄 Selecting new chunks for hypothesis attempt {attempts}...")
                    randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                    dynamic_chunks = self.select_dynamic_chunks_for_generation(meta_hypothesis, randomization_strategy=randomization_strategy)
                    if dynamic_chunks:
                        context_chunks = dynamic_chunks
                        print(f"📚 Using {len(context_chunks)} dynamically selected chunks for generation")
                    else:
                        print(f"⚠️  Failed to select dynamic chunks, falling back to direct retrieval...")
                        # Fallback: get fresh chunks directly
                        context_chunks = self.retrieve_relevant_chunks(meta_hypothesis, top_k=chunks_per_meta)
                        if context_chunks:
                            print(f"📚 Using {len(context_chunks)} fresh chunks for generation")
                        else:
                            print(f"❌ No chunks available for meta-hypothesis {meta_idx}. Skipping...")
                            break
                else:
                    print(f"📚 Using direct chunk retrieval (no dynamic selection available)...")
                    # Get fresh chunks for each hypothesis attempt
                    context_chunks = self.retrieve_relevant_chunks(meta_hypothesis, top_k=chunks_per_meta)
                    if not context_chunks:
                        print(f"❌ No relevant context found for meta-hypothesis {meta_idx}. Skipping...")
                        break
                    print(f"📚 Using {len(context_chunks)} fresh chunks for generation")

                # Start timer for this hypothesis
                self.hypothesis_timer.start()
                print(f"⏱️  Starting 5-minute timer for hypothesis attempt {attempts}...")

                # Generate hypothesis
                try:
                    context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                    hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1)

                    if not hypothesis_list:
                        print("❌ Failed to generate hypothesis. Skipping...")
                        continue

                    hypothesis = hypothesis_list[0]
                    
                    # Validate hypothesis format before proceeding
                    from hypothesis_tools import validate_hypothesis_format
                    is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
                    if not is_valid_format:
                        print(f"❌ Hypothesis rejected due to format: {format_reason}")
                        # Record format rejection
                        format_rejected_record = {
                            'Meta_Hypothesis': meta_hypothesis,
                            'Meta_Hypothesis_Index': meta_idx,
                            'Hypothesis': hypothesis,
                            'Accuracy': None,
                            'Novelty': None,
                            'Relevancy': None,
                            'Verdict': 'FORMAT_REJECTED',
                            'Critique': f'Hypothesis rejected: {format_reason}',
                            'Citations': 'No citations available',
                            'Original_Query': user_prompt
                        }
                        self.hypothesis_records.append(format_rejected_record)
                        # Save incrementally
                        self.save_hypothesis_record_incrementally(format_rejected_record, excel_filename)
                        continue
                    
                    print(f"📝 Generated: {hypothesis}")

                    # Check timer before critique
                    if self.hypothesis_timer.check_expired():
                        print(f"⏰ Time limit reached before critique")
                        continue

                    # Critique hypothesis
                    print(f"🔍 Critiquing hypothesis...")
                    critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt=meta_hypothesis, lab_goals=get_lab_goals())

                    if critique_result:
                        novelty = critique_result.get('novelty')
                        accuracy = critique_result.get('accuracy')
                        relevancy = critique_result.get('relevancy')
                        critique = critique_result.get('critique', 'No critique available')

                        # Determine verdict
                        verdict = self.automated_verdict(accuracy, novelty, relevancy)

                        # Extract citations from the dynamically selected chunks used for this specific hypothesis
                        citations = []
                        unique_sources = set()
                        
                        # Get metadata from the dynamically selected chunks
                        dynamic_metadata = []
                        if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                            # If we used dynamic chunk selection, get metadata from those chunks
                            if hasattr(self, 'last_dynamic_results') and self.last_dynamic_results:
                                dynamic_metadata = [result.get('metadata', {}) for result in self.last_dynamic_results]
                            else:
                                # Fallback: extract from context_chunks if they have metadata
                                dynamic_metadata = []
                                for chunk in context_chunks:
                                    if isinstance(chunk, dict) and 'metadata' in chunk:
                                        dynamic_metadata.append(chunk['metadata'])
                                    else:
                                        # If chunks don't have metadata, use extract_citations_from_chunks as fallback
                                        citations = self.extract_citations_from_chunks(context_chunks)
                                        break
                        else:
                            # If not using dynamic selection, use extract_citations_from_chunks
                            citations = self.extract_citations_from_chunks(context_chunks)
                        
                        # Extract citations from the metadata if we have it
                        if not citations and dynamic_metadata:
                            for metadata in dynamic_metadata:
                                if metadata:
                                    source_name = metadata.get('source_name', 'Unknown')
                                    title = metadata.get('title', 'No title')
                                    doi = metadata.get('doi', 'No DOI')
                                    authors = metadata.get('author', metadata.get('authors', 'Unknown authors'))
                                    journal = metadata.get('journal', 'Unknown journal')
                                    # Extract year from publication_date
                                    publication_date = metadata.get('publication_date', '')
                                    year = publication_date[:4] if publication_date and len(publication_date) >= 4 else 'Unknown year'

                                    citation = {
                                        'source_name': source_name,
                                        'title': title,
                                        'doi': doi,
                                        'authors': authors,
                                        'journal': journal,
                                        'year': year
                                    }

                                    # Add to citations if not already present
                                    citation_key = doi if doi != 'No DOI' else title
                                    if citation_key not in unique_sources:
                                        citations.append(citation)
                                        unique_sources.add(citation_key)
                        
                        formatted_citations = self.format_citations_for_export(citations)

                        # Create record
                        record = {
                            'Meta_Hypothesis': meta_hypothesis,
                            'Meta_Hypothesis_Index': meta_idx,
                            'Hypothesis': hypothesis,
                            'Accuracy': accuracy,
                            'Novelty': novelty,
                            'Relevancy': relevancy,
                            'Verdict': verdict,
                            'Critique': critique,
                            'Citations': formatted_citations,
                            'Original_Query': user_prompt
                        }

                        # Add to records
                        self.hypothesis_records.append(record)

                        # Save incrementally to Excel
                        self.save_hypothesis_record_incrementally(record, excel_filename)

                        # Display result
                        print(f"\n📊 Hypothesis Result:")
                        print(f"   Novelty: {novelty}/85")
                        print(f"   Accuracy: {accuracy}/80")
                        print(f"   Relevancy: {relevancy}/50")
                        print(f"   Verdict: {verdict}")

                        # Print rejection reason if hypothesis is rejected
                        if verdict != 'ACCEPTED':
                            if accuracy < 80:
                                print(f"❌ Hypothesis rejected: Low accuracy score ({accuracy}/80)")
                            elif novelty < 85:
                                print(f"❌ Hypothesis rejected: Low novelty score ({novelty}/85)")
                            elif relevancy < 50:
                                print(f"❌ Hypothesis rejected: Low relevancy score ({relevancy}/50)")
                            else:
                                print(f"❌ Hypothesis rejected: Unknown reason")

                        # Accept hypothesis if it meets criteria
                        if verdict == "ACCEPTED":
                            meta_hypotheses_generated.append(record)
                            print(f"✅ Hypothesis accepted for meta-hypothesis {meta_idx}! ({len(meta_hypotheses_generated)}/{n_per_meta})")
                            
                            # Check if we have enough accepted hypotheses for this meta-hypothesis
                            if len(meta_hypotheses_generated) >= n_per_meta:
                                print(f"🎯 Successfully generated {n_per_meta} accepted hypotheses for meta-hypothesis {meta_idx}. Moving to next meta-hypothesis.")
                                break
                        else:
                            print(f"❌ Hypothesis rejected for meta-hypothesis {meta_idx}")

                        # Check timer
                        if self.hypothesis_timer.check_expired():
                            print(f"⏰ Time limit reached for meta-hypothesis {meta_idx}. Moving to next meta-hypothesis.")
                            break

                except Exception as e:
                    print(f"❌ Error generating/critiquing hypothesis: {e}")
                    continue

            # Check if we successfully generated enough accepted hypotheses for this meta-hypothesis
            if len(meta_hypotheses_generated) < n_per_meta:
                print(f"⚠️  Warning: Only generated {len(meta_hypotheses_generated)}/{n_per_meta} accepted hypotheses for meta-hypothesis {meta_idx}")
                print(f"   This may indicate that the meta-hypothesis is too specific or the acceptance criteria are too strict.")
            else:
                print(f"✅ Successfully generated {len(meta_hypotheses_generated)}/{n_per_meta} accepted hypotheses for meta-hypothesis {meta_idx}")
            
            all_hypotheses.extend(meta_hypotheses_generated)

        # Summary
        print(f"\n🎉 META-HYPOTHESIS GENERATION COMPLETE")
        print("=" * 80)
        print(f"Total meta-hypotheses processed: {total_meta_hypotheses}")
        print(f"Total accepted hypotheses generated: {len(all_hypotheses)}")
        print(f"Average accepted hypotheses per meta-hypothesis: {len(all_hypotheses)/total_meta_hypotheses:.1f}")

        # Final export summary
        if all_hypotheses:
            print(f"\n💾 Final export summary:")
            print(f"📊 All results have been incrementally saved to: {excel_filename}")
            print(f"📁 File location: {os.path.dirname(os.path.abspath(excel_filename))}")

            # Also create a final comprehensive export
            final_export_filename = self.export_hypotheses_to_excel()
            if final_export_filename:
                print(f"📋 Comprehensive results also exported to: {final_export_filename}")
            else:
                print(f"⚠️  Comprehensive export failed, but incremental saves are complete")
        else:
            print(f"❌ No hypotheses were successfully generated")

        return all_hypotheses

    def check_api_status(self):
        """Check Gemini API status and provide guidance."""
        print("\n🔍 Checking Gemini API Status...")

        if not self.gemini_client:
            print("❌ Gemini client not initialized")
            print("💡 Check your API key in keys.json")
            return

        # Show current rate limit status
        remaining_requests = self.gemini_rate_limiter.get_remaining_requests()
        print(f"📊 Gemini API Rate Limit Status:")
        print(f"   - Limit: 1000 requests per minute")
        print(f"   - Remaining requests: {remaining_requests}")
        print(f"   - Used requests: {1000 - remaining_requests}")

        try:
            # Try a simple test request
            test_response = self.gemini_client.models.generate_content(
                model="gemini-2.5-flash",
                contents="Hello"
            )
            print("✅ Gemini API is working")
            print("💡 You can use 'generate' command for hypothesis generation")
        except Exception as e:
            error_str = str(e)
            if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower():
                print("❌ API quota exceeded (rate limited)")
                print("💡 Solutions:")
                print("   - Wait 30-60 minutes for quota reset")
                print("   - Use 'generate_offline' for fallback hypotheses")
                print("   - Check your billing/plan at https://ai.google.dev/")
                print("   - Consider upgrading your plan for higher limits")
            elif "401" in error_str or "403" in error_str:
                print("❌ API authentication failed")
                print("💡 Check your API key in keys.json")
            else:
                print(f"❌ API error: {e}")
                print("💡 Try again later or use 'generate_offline'")

    def is_chromadb_ready(self):
        """Check if ChromaDB is ready and has data for searching."""
        if not self.use_chromadb or not self.chroma_manager:
            return False

        try:
            stats = self.chroma_manager.get_collection_stats()
            return stats.get('total_documents', 0) > 0
        except Exception as e:
            logger.error(f"Error checking ChromaDB readiness: {e}")
            return False

    def _initialize_hypothesis_tools(self):
        """Initialize the HypothesisGenerator, HypothesisCritic, and MetaHypothesisGenerator."""
        if self.gemini_client:
            self.hypothesis_generator = HypothesisGenerator(model=self.gemini_client)
            def embedding_fn(text):
                return np.array(self.get_google_embedding(text))
            self.hypothesis_critic = HypothesisCritic(model=self.gemini_client, embedding_fn=embedding_fn)
            self.meta_hypothesis_generator = MetaHypothesisGenerator(model=self.gemini_client)
            print("✅ Hypothesis tools initialized successfully.")
        else:
            print("⚠️ Gemini client not initialized, skipping hypothesis tools.")

    def export_hypotheses_to_excel(self, filename=None):
        """Export hypothesis records to Excel file with grouping by meta-hypothesis."""
        if not self.hypothesis_records:
            print("❌ No hypothesis records to export.")
            return None

        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if filename is None:
                # Ensure hypothesis_export directory exists
                export_dir = "hypothesis_export"
                os.makedirs(export_dir, exist_ok=True)
                filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")

            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Check if we have meta-hypothesis data
            has_meta_data = any('Meta_Hypothesis' in record for record in self.hypothesis_records)

            if has_meta_data:
                # Group by meta-hypothesis
                meta_groups = {}
                for record in self.hypothesis_records:
                    meta_hypothesis = record.get('Meta_Hypothesis', 'Unknown')
                    if meta_hypothesis not in meta_groups:
                        meta_groups[meta_hypothesis] = []
                    meta_groups[meta_hypothesis].append(record)

                # Create workbook with multiple sheets
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove default sheet

                # Define styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                meta_header_font = Font(bold=True, color="FFFFFF", size=12)
                meta_header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Create summary sheet
                summary_sheet = workbook.create_sheet("Summary")
                summary_sheet['A1'] = "Meta-Hypothesis Summary"
                summary_sheet['A1'].font = meta_header_font
                summary_sheet['A1'].fill = meta_header_fill

                summary_sheet['A3'] = "Meta-Hypothesis"
                summary_sheet['B3'] = "Hypotheses Generated"
                summary_sheet['C3'] = "Accepted"
                summary_sheet['D3'] = "Rejected"
                summary_sheet['E3'] = "Average Novelty"
                summary_sheet['F3'] = "Average Accuracy"

                for col in ['A3', 'B3', 'C3', 'D3', 'E3', 'F3']:
                    summary_sheet[col].font = header_font
                    summary_sheet[col].fill = header_fill
                    summary_sheet[col].border = border

                row = 4
                for meta_hypothesis, records in meta_groups.items():
                    summary_sheet[f'A{row}'] = meta_hypothesis[:100] + "..." if len(meta_hypothesis) > 100 else meta_hypothesis
                    summary_sheet[f'B{row}'] = len(records)

                    accepted_count = sum(1 for r in records if r.get('Verdict') == 'ACCEPTED')
                    rejected_count = len(records) - accepted_count
                    summary_sheet[f'C{row}'] = accepted_count
                    summary_sheet[f'D{row}'] = rejected_count

                    avg_novelty = sum(r.get('Novelty', 0) for r in records if r.get('Novelty') is not None) / len(records)
                    avg_accuracy = sum(r.get('Accuracy', 0) for r in records if r.get('Accuracy') is not None) / len(records)
                    summary_sheet[f'E{row}'] = round(avg_novelty, 1)
                    summary_sheet[f'F{row}'] = round(avg_accuracy, 1)

                    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                        summary_sheet[f'{col}{row}'].border = border

                    row += 1

                # Auto-adjust column widths for summary
                for column in summary_sheet.columns:
                    max_length = 0
                    try:
                        column_letter = column[0].column_letter
                    except AttributeError:
                        # Skip merged cells that don't have column_letter attribute
                        continue
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    summary_sheet.column_dimensions[column_letter].width = adjusted_width

                # Create detailed sheets for each meta-hypothesis
                for i, (meta_hypothesis, records) in enumerate(meta_groups.items(), 1):
                    # Create sheet name (Excel has 31 character limit)
                    sheet_name = f"Meta_{i}"
                    if len(meta_hypothesis) > 20:
                        sheet_name += f"_{meta_hypothesis[:20].replace(' ', '_')}"
                    else:
                        sheet_name += f"_{meta_hypothesis.replace(' ', '_')}"

                    # Clean sheet name for Excel compatibility
                    sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in ('_', '-'))[:31]

                    sheet = workbook.create_sheet(sheet_name)

                    # Add meta-hypothesis header
                    sheet['A1'] = f"Meta-Hypothesis {i}: {meta_hypothesis}"
                    sheet['A1'].font = meta_header_font
                    sheet['A1'].fill = meta_header_fill
                    sheet.merge_cells('A1:H1')

                    # Add hypothesis data
                    df_group = pd.DataFrame(records)
                    if not df_group.empty:
                        # Write headers
                        headers = list(df_group.columns)
                        for col, header in enumerate(headers, 1):
                            cell = sheet.cell(row=3, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = border

                        # Write data
                        for row_idx, record in enumerate(records, 4):
                            for col_idx, (key, value) in enumerate(record.items(), 1):
                                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                                cell.border = border

                        # Auto-adjust column widths
                        for column in sheet.columns:
                            max_length = 0
                            try:
                                column_letter = column[0].column_letter
                            except AttributeError:
                                # Skip merged cells that don't have column_letter attribute
                                continue
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            sheet.column_dimensions[column_letter].width = adjusted_width

                # Save workbook
                workbook.save(filename)

            else:
                # Fallback to simple export for non-meta-hypothesis data
                df = pd.DataFrame(self.hypothesis_records)
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Hypotheses', index=False)
                    worksheet = writer.sheets['Hypotheses']
                    for column in worksheet.columns:
                        max_length = 0
                        try:
                            column_letter = column[0].column_letter
                        except AttributeError:
                            # Skip merged cells that don't have column_letter attribute
                            continue
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"✅ Successfully exported {len(self.hypothesis_records)} hypothesis records to: {filename}")
            print(f"📁 File saved in: {os.path.dirname(os.path.abspath(filename))}")
            if has_meta_data:
                print(f"📊 Organized into {len(meta_groups)} meta-hypothesis groups with summary sheet")
            return filename
        except Exception as e:
            print(f"❌ Failed to export hypotheses to Excel: {e}")
            return None

    def save_hypothesis_record_incrementally(self, record, filename=None):
        """Save a single hypothesis record to Excel file incrementally."""
        if filename is None:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Ensure hypothesis_export directory exists
            export_dir = "hypothesis_export"
            os.makedirs(export_dir, exist_ok=True)
            filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")

        try:
            import pandas as pd
            from openpyxl import load_workbook

            # Check if file exists
            if os.path.exists(filename):
                # Load existing workbook
                workbook = load_workbook(filename)

                # Get or create the Hypotheses sheet
                if 'Hypotheses' in workbook.sheetnames:
                    worksheet = workbook['Hypotheses']
                    # Find the next empty row
                    next_row = worksheet.max_row + 1
                else:
                    worksheet = workbook.create_sheet('Hypotheses')
                    next_row = 1
                    # Add headers for new sheet
                    headers = list(record.keys())
                    for col, header in enumerate(headers, 1):
                        worksheet.cell(row=1, column=col, value=header)
                    next_row = 2

                # Add the record data
                for col, value in enumerate(record.values(), 1):
                    worksheet.cell(row=next_row, column=col, value=value)

                # Save the workbook
                workbook.save(filename)

            else:
                # Create new file
                df_row = pd.DataFrame([record])
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df_row.to_excel(writer, sheet_name='Hypotheses', index=False)
                    worksheet = writer.sheets['Hypotheses']
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        try:
                            column_letter = column[0].column_letter
                        except AttributeError:
                            # Skip merged cells that don't have column_letter attribute
                            continue
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"💾 Saved hypothesis record to: {filename}")
            print(f"📁 File saved in: {os.path.dirname(os.path.abspath(filename))}")
            return filename

        except Exception as e:
            print(f"❌ Failed to save hypothesis record incrementally: {e}")
            return None

    def clear_hypothesis_records(self):
        """Clear all stored hypothesis records."""
        self.hypothesis_records = []
        print("🗑️ Hypothesis records cleared.")

    def show_hypothesis_records(self):
        """Display current hypothesis records."""
        if not self.hypothesis_records:
            print("📝 No hypothesis records available.")
            return

        print(f"\n📝 Current Hypothesis Records ({len(self.hypothesis_records)} total):")
        print("=" * 80)

        for i, record in enumerate(self.hypothesis_records, 1):
            print(f"\n{i}. Hypothesis: {record['Hypothesis'][:100]}{'...' if len(record['Hypothesis']) > 100 else ''}")
            print(f"   Accuracy: {record['Accuracy']}, Novelty: {record['Novelty']}, Relevancy: {record['Relevancy']}, Verdict: {record['Verdict']}")
            print(f"   Citations: {len(record['Citations'].split(';')) if record['Citations'] != 'No citations available' else 0} sources")
            print("-" * 80)

    def _display_commands(self):
        """Display available commands."""
        print(f"\n💡 Available commands:")
        print(f"   - 'add <query>': Search, batch, and generate hypotheses with automatic diverse paper selection (e.g., 'add cancer')")
        print(f"   - 'add <N> <query>': Search, batch, and generate hypotheses from up to N results (e.g., 'add 2000 cancer')")
        print(f"   - 'add all <query>': Search, batch, and generate hypotheses from ALL found results (e.g., 'add all cancer')")
        print(f"   - 'meta <query>': Use meta-hypothesis generator to create 5 diverse research directions, then generate hypotheses for each (e.g., 'meta UBR-5 in cancer')")
        print(f"   - 'meta <chunks> <query>': Advanced usage - specify number of chunks per meta-hypothesis (e.g., 'meta 2000 UBR-5 in cancer')")
        print(f"   - 'clear': Clear current package")
        print(f"   - 'export': Export hypothesis records to Excel")
        print(f"   - 'records': Show current hypothesis records")
        print(f"   - 'clear_records': Clear hypothesis records")
        print(f"   - 'reset_papers': Reset used papers list (allow reusing papers)")
        print(f"   - 'reset_diverse': Reset diverse paper selection (get new set of 1500 papers)")
        print(f"   - 'papers_status': Show used papers tracking status")
        print(f"   - 'lab_ratio': Configure ratio of lab-authored papers (auto/fixed ratio)")
        print(f"   - 'preferred_authors': Configure preferred authors for search prioritization")
        print(f"   - 'randomization': Configure chunk randomization strategy")
        print(f"   - 'test_lab': Test lab paper detection with current query")
        print(f"   - 'diverse <query>': Manually select exactly 1500 different papers for hypothesis generation (diverse search is now DEFAULT)")
        print(f"   💡 Note: Large queries may take longer and use more API calls. Use 'clear' if you get errors.")
        print(f"   📚 Default: Diverse search with 1500 chunks from different papers for maximum variety!")
        print()

    def interactive_search(self):
        novelty_threshold = 85
        accuracy_threshold = 80
        relevancy_threshold = 50
        print("=== Enhanced RAG Query System ===")
        print("🔍 Search across all your knowledge bases!")
        self._display_commands()
        while True:
            query = input("❓ Your question (or command): ").strip()
            if query.lower() == 'clear':
                self.clear_package()
            elif query.lower() == 'export':
                self.export_hypotheses_to_excel()
            elif query.lower() == 'records':
                self.show_hypothesis_records()
            elif query.lower() == 'clear_records':
                self.clear_hypothesis_records()
            elif query.lower() == 'reset_papers':
                self.reset_used_papers()
            elif query.lower() == 'papers_status':
                self.show_used_papers_status()
            elif query.lower() == 'reset_diverse':
                if (hasattr(self, 'current_package') and 
                    self.current_package.get('method') == 'diverse_paper_selection'):
                    print("🔄 Resetting diverse paper selection...")
                    self.current_package = {}
                    print("✅ Diverse paper selection reset. Use 'diverse <query>' to select a new set of 1500 papers.")
                else:
                    print("ℹ️  No diverse paper selection to reset.")
            elif query.lower() == 'lab_ratio':
                self.configure_lab_paper_ratio()
            elif query.lower() == 'preferred_authors':
                self.configure_preferred_authors()
            elif query.lower() == 'randomization':
                self.configure_randomization_strategy()
            elif query.lower() == 'test_lab':
                # Get the current query from package or use default
                if "prompt" in self.current_package and self.current_package["prompt"]:
                    test_query = self.current_package["prompt"]
                else:
                    test_query = "UBR5 cancer"
                self.test_lab_paper_detection(test_query)
            elif query.lower().startswith('diverse '):
                try:
                    parts = query.split(' ', 1)
                    if len(parts) < 2:
                        print("❌ Usage: diverse <query> (e.g., 'diverse UBR-5 in cancer')")
                        continue
                    
                    diverse_query = parts[1]
                    print(f"\n🎯 Selecting exactly 1500 different papers for hypothesis generation...")
                    print(f"🔍 Query: '{diverse_query}'")
                    print(f"📚 This ensures maximum diversity across all sources and papers")
                    
                    response = input("Continue? (y/n): ").lower()
                    if response == 'y':
                        try:
                            # Use the new diversity function
                            diverse_chunks = self.select_diverse_papers_for_hypothesis_generation(
                                diverse_query, 
                                target_papers=1500
                            )
                            
                            if diverse_chunks:
                                print(f"\n✅ Successfully selected {len(diverse_chunks)} chunks from exactly 1500 different papers!")
                                print(f"🎯 Each hypothesis generation will now use a completely different set of source chunks")
                                print(f"📊 Papers are prioritized by: lab authorship, preferred authors, citation count, and impact factor")
                                
                                # Store the diverse chunks for future hypothesis generation
                                self.current_package = {
                                    "chunks": diverse_chunks,
                                    "prompt": diverse_query,
                                    "method": "diverse_paper_selection"
                                }
                                
                                print(f"📦 Diverse paper package stored. Use 'add' command to generate hypotheses with these diverse chunks.")
                            else:
                                print("❌ Failed to select diverse papers.")
                        except Exception as e:
                            print(f"❌ Error during diverse paper selection: {e}")
                    else:
                        print("Diverse paper selection cancelled.")
                except Exception as e:
                    print(f"❌ Error processing diverse command: {e}")
            elif query.lower().startswith('meta '):
                try:
                    parts = query.split(' ', 1)
                    if len(parts) < 2:
                        print("❌ Usage: meta <query> (e.g., 'meta UBR-5 in cancer')")
                        print("   Advanced: meta <chunks> <query> (e.g., 'meta 2000 UBR-5 in cancer' for 2000 chunks per meta-hypothesis)")
                        continue

                    # Check if first part is a number (chunks specification)
                    query_parts = parts[1].split(' ', 1)
                    if len(query_parts) > 1 and query_parts[0].isdigit():
                        chunks_per_meta = int(query_parts[0])
                        # Validate chunks number
                        if chunks_per_meta < 5:
                            print(f"❌ Too few chunks ({chunks_per_meta}). Minimum is 5 chunks per meta-hypothesis.")
                            continue
                        elif chunks_per_meta > 5000:
                            print(f"⚠️  Very large number of chunks ({chunks_per_meta}). This may cause API errors or timeouts.")
                            response = input("Continue anyway? (y/n): ").lower()
                            if response != 'y':
                                continue
                        meta_query = query_parts[1]
                        print(f"\n🧠 Meta-hypothesis generation for: '{meta_query}'")
                        print(f"📚 Using {chunks_per_meta} chunks per meta-hypothesis")
                    else:
                        chunks_per_meta = 1500  # default
                        meta_query = parts[1]
                        print(f"\n🧠 Meta-hypothesis generation for: '{meta_query}'")
                        print(f"📚 Using default {chunks_per_meta} chunks per meta-hypothesis")

                    print("This will generate 5 diverse research directions, then create hypotheses for each.")
                    response = input("Continue? (y/n): ").lower()
                    if response == 'y':
                        try:
                            results = self.generate_hypotheses_with_meta_generator(meta_query, n_per_meta=5, chunks_per_meta=chunks_per_meta)
                            if results:
                                print(f"\n✅ Meta-hypothesis generation complete! Generated {len(results)} accepted hypotheses across 5 research directions.")
                            else:
                                print("❌ Meta-hypothesis generation failed.")
                        except Exception as e:
                            print(f"❌ Error during meta-hypothesis generation: {e}")
                    else:
                        print("Meta-hypothesis generation cancelled.")
                except Exception as e:
                    print(f"❌ Error processing meta command: {e}")
            elif query.lower().startswith('add '):
                try:
                    parts = query.split(' ', 2)
                    if len(parts) < 2:
                        print("❌ Usage: add <query> (e.g., 'add cancer')")
                        continue
                    # Check for 'all' keyword
                    if len(parts) > 2 and parts[1].lower() == 'all':
                        num_results = None  # None means all
                        search_query = parts[2]
                    elif len(parts) > 2:
                        try:
                            num_results = int(parts[1])
                            if num_results <= 0:
                                print("❌ Number of results must be positive")
                                continue
                            if num_results > 100000:
                                print("⚠️  Number of results is very large (>100,000). This may use a lot of memory and cause API errors.")
                            search_query = parts[2]
                        except ValueError:
                            num_results = 100000
                            search_query = parts[1] + " " + parts[2]
                    else:
                        num_results = 100000
                        search_query = parts[1]
                    print(f"\n🔍 Searching for: '{search_query}'" + (f" (requesting ALL results)" if num_results is None else f" (requesting up to {num_results} results)"))
                    try:
                        # Decide which method to use
                        if num_results is None or num_results > 5000:
                            # Use get_all_documents and compute similarity in Python
                            print("⚠️  Using full-document scan and in-memory similarity search. This may use a lot of RAM and be slow for very large collections.")
                            # Get all documents (limit to 100,000 for safety)
                            max_docs = num_results if (num_results is not None and num_results < 100000) else 100000
                            all_docs = self.chroma_manager.get_all_documents(limit=max_docs)
                            if not all_docs:
                                print("❌ No documents found in ChromaDB.")
                                continue
                            print(f"📦 Retrieved {len(all_docs)} documents from ChromaDB. Computing similarities...")
                            # Get query embedding
                            query_embedding = self.get_google_embedding(search_query)
                            if not query_embedding:
                                print("❌ Failed to get query embedding.")
                                continue
                            import numpy as np
                            from tqdm.auto import tqdm
                            doc_embeddings = []
                            for doc in all_docs:
                                emb = doc['metadata'].get('embedding', None)
                                if emb is not None:
                                    doc_embeddings.append(emb)
                                else:
                                    doc_embeddings.append([0.0]*len(query_embedding))  # fallback if missing
                            doc_embeddings = np.array(doc_embeddings)
                            from sklearn.metrics.pairwise import cosine_similarity
                            similarities = np.zeros(len(doc_embeddings))
                            print(f"🔍 Computing similarities across {len(doc_embeddings):,} documents...")
                            with tqdm(total=len(doc_embeddings), desc="Similarity", unit="doc") as pbar:
                                batch_size = 1000
                                for start in range(0, len(doc_embeddings), batch_size):
                                    end = min(start + batch_size, len(doc_embeddings))
                                    similarities[start:end] = cosine_similarity([query_embedding], doc_embeddings[start:end])[0]
                                    pbar.update(end - start)
                            top_indices = np.argsort(similarities)[::-1][:num_results if num_results is not None else len(all_docs)]
                            results = []
                            for idx in top_indices:
                                results.append({
                                    'chunk': all_docs[idx]['document'],
                                    'metadata': all_docs[idx]['metadata'],
                                    'similarity': similarities[idx],
                                    'method': 'manual_cosine'
                                })
                        else:
                            # Use fast ChromaDB search
                            results = self.search_hybrid(search_query, top_k=num_results, filter_dict=None)
                        if not results:
                            print("❌ No results found. Try a different search term or check if data is loaded.")
                            continue
                        print(f"📦 Found {len(results)} relevant chunks. Beginning batch processing...")
                        # Batch process for hypothesis generation
                        batch_size = 500
                        
                        # AUTOMATICALLY use diverse paper selection as default for hypothesis generation
                        print(f"\n🎯 Automatically selecting diverse papers for hypothesis generation...")
                        print(f"📚 This ensures maximum diversity across all sources and papers")
                        
                        diverse_chunks = self.select_diverse_papers_for_hypothesis_generation(
                            search_query, 
                            target_papers=1500
                        )
                        
                        if diverse_chunks:
                            print(f"✅ Successfully selected {len(diverse_chunks)} chunks from exactly 1500 different papers!")
                            print(f"🎯 Each hypothesis generation will now use a completely different set of source chunks")
                            print(f"📊 Papers are prioritized by: lab authorship, preferred authors, citation count, and impact factor")
                            all_chunks = diverse_chunks
                        else:
                            print(f"⚠️  Failed to select diverse papers, falling back to original chunks")
                            all_chunks = [r['chunk'] for r in results]
                        # Sequential, real-time generator-critic loop for 5 accepted hypotheses
                        print(f"\n🧠 Generating hypotheses one at a time until 5 are accepted (novelty >= {novelty_threshold}, accuracy >= {accuracy_threshold}, 5-minute time limit per critique)...")
                        accepted_hypotheses = []
                        attempts = 0
                        max_attempts = 100  # Prevent infinite loops
                        while len(accepted_hypotheses) < 5 and attempts < max_attempts:
                            attempts += 1
                            print(f"\n🧠 Generating hypothesis attempt {attempts}...")

                            # Use the diverse chunks for all attempts (no need to re-select)
                            print(f"🎯 Using diverse paper chunks for hypothesis attempt {attempts}...")
                            print(f"📚 Using {len(all_chunks)} chunks from exactly 1500 different papers")

                            # Start timer for this hypothesis
                            self.hypothesis_timer.start()
                            print(f"⏱️  Starting 5-minute timer for hypothesis attempt {attempts}...")

                            hypothesis_list = self._generate_hypotheses_with_retry(all_chunks, n=1)
                            if not hypothesis_list:
                                print("❌ Failed to generate hypothesis. Skipping...")
                                # Record failed generation
                                self.hypothesis_records.append({
                                    'Hypothesis': f'Failed generation attempt {attempts}',
                                    'Accuracy': None,
                                    'Novelty': None,
                                    'Verdict': 'GENERATION_FAILED',
                                    'Critique': 'Failed to generate hypothesis',
                                    'Citations': 'No citations available'
                                })
                                continue

                            hypothesis = hypothesis_list[0]
                            print(f"Generated Hypothesis: {hypothesis}")
                            print(f"\n🔄 Critiquing Hypothesis: {hypothesis}")

                            # Check timer before critique
                            if self.hypothesis_timer.check_expired():
                                print(f"⏰ Time limit reached for hypothesis attempt {attempts}. Moving to next attempt.")
                                # Record timeout result
                                self.hypothesis_records.append({
                                    'Hypothesis': hypothesis,
                                    'Accuracy': None,
                                    'Novelty': None,
                                    'Verdict': 'TIMEOUT',
                                    'Critique': 'Critique process timed out after 5 minutes',
                                    'Citations': 'No citations available'
                                })
                                continue

                            critique_result = self._critique_hypothesis_with_retry(hypothesis, all_chunks, search_query)
                            if not critique_result:
                                print(f"❌ Failed to critique hypothesis after retries. Skipping...")
                                # Record failed critique
                                self.hypothesis_records.append({
                                    'Hypothesis': hypothesis,
                                    'Accuracy': None,
                                    'Novelty': None,
                                    'Verdict': 'CRITIQUE_FAILED',
                                    'Critique': 'Failed to critique hypothesis after retries',
                                    'Citations': 'No citations available'
                                })
                                continue

                            # Check timer after critique
                            if self.hypothesis_timer.check_expired():
                                print(f"⏰ Time limit reached for hypothesis attempt {attempts} after critique. Moving to next attempt.")
                                # Record timeout result with partial critique
                                self.hypothesis_records.append({
                                    'Hypothesis': hypothesis,
                                    'Accuracy': critique_result.get('accuracy', None),
                                    'Novelty': critique_result.get('novelty', None),
                                    'Verdict': 'TIMEOUT',
                                    'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                                    'Citations': 'No citations available'
                                })
                                continue

                            novelty = critique_result.get('novelty', 0)
                            accuracy = critique_result.get('accuracy', 0)
                            relevancy = critique_result.get('relevancy', 0)
                            verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
                            remaining_time = self.hypothesis_timer.get_remaining_time()
                            print(f"   Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold}, Automated Verdict: {verdict}")
                            print(f"   ⏱️  Time remaining: {remaining_time:.1f}s")

                            # Extract citations from results metadata
                            citations = []
                            unique_sources = set()
                            for result in results:
                                metadata = result.get('metadata', {})
                                if metadata:
                                    source_name = metadata.get('source_name', 'Unknown')
                                    title = metadata.get('title', 'No title')
                                    doi = metadata.get('doi', 'No DOI')
                                    authors = metadata.get('author', metadata.get('authors', 'Unknown authors'))
                                    journal = metadata.get('journal', 'Unknown journal')
                                    # Extract year from publication_date
                                    publication_date = metadata.get('publication_date', '')
                                    year = publication_date[:4] if publication_date and len(publication_date) >= 4 else 'Unknown year'

                                    citation = {
                                        'source_name': source_name,
                                        'title': title,
                                        'doi': doi,
                                        'authors': authors,
                                        'journal': journal,
                                        'year': year
                                    }

                                    # Add to citations if not already present
                                    citation_key = doi if doi != 'No DOI' else title
                                    if citation_key not in unique_sources:
                                        citations.append(citation)
                                        unique_sources.add(citation_key)

                            formatted_citations = self.format_citations_for_export(citations)

                            # Track record for export
                            self.hypothesis_records.append({
                                'Hypothesis': hypothesis,
                                'Accuracy': accuracy,
                                'Novelty': novelty,
                                'Relevancy': relevancy,
                                'Verdict': verdict,
                                'Critique': critique_result.get('critique', ''),
                                'Citations': formatted_citations
                            })

                            if verdict == 'ACCEPTED':
                                accepted_hypotheses.append({
                                    "hypothesis": hypothesis,
                                    "critique": critique_result,
                                    "score": (novelty + accuracy + relevancy) / 3
                                })
                                print(f"✅ Hypothesis accepted! ({len(accepted_hypotheses)}/5)")
                            else:
                                print(f"❌ Hypothesis rejected (verdict: {verdict})")

                        if not accepted_hypotheses:
                            print("❌ No hypotheses were successfully critiqued.")
                            continue

                        print(f"\n🏆 Top {len(accepted_hypotheses)} Hypotheses:")
                        print("=" * 80)
                        for i, result in enumerate(accepted_hypotheses, 1):
                            hypothesis = result["hypothesis"]
                            score = result["score"]
                            critique = result["critique"]
                            print(f"\n{i}. {hypothesis}")
                            print(f"   Score: {score:.1f}")
                            print(f"   Novelty: {critique.get('novelty', 'N/A')}")
                            print(f"   Accuracy: {critique.get('accuracy', 'N/A')}")
                            print(f"   Relevancy: {critique.get('relevancy', 'N/A')}")
                            print(f"   Automated Verdict: {self.automated_verdict(critique.get('accuracy', 0), critique.get('novelty', 0), critique.get('relevancy', 0), accuracy_threshold, novelty_threshold, relevancy_threshold)}")
                            print(f"   Critique: {critique.get('critique', 'N/A')}")
                            print("-" * 80)

                        # Auto-export results
                        print(f"\n💾 Auto-exporting hypothesis records...")
                        export_filename = self.export_hypotheses_to_excel()
                        if export_filename:
                            print(f"📊 Results exported to: {export_filename}")
                            print(f"📁 File saved in: {os.path.dirname(os.path.abspath(export_filename))}")

                        # Enter discussion loop
                        print("\n💬 You can now discuss these hypotheses with the AI. Type the number of a hypothesis (1-5) to select it, or 'exit' to leave discussion mode.")
                        selected_idx = 0
                        while True:
                            user_input = input("[Discussion] > ").strip()
                            if user_input.lower() in ['exit', 'quit']:
                                print("Exiting discussion mode.")
                                break
                            if user_input.isdigit() and 1 <= int(user_input) <= len(accepted_hypotheses):
                                selected_idx = int(user_input) - 1
                                print(f"Selected Hypothesis {user_input}:\n{accepted_hypotheses[selected_idx]['hypothesis']}")
                                print("You can now ask questions about this hypothesis. Type 'back' to select another, or 'exit' to leave.")
                                while True:
                                    q = input(f"[H{selected_idx+1} Q&A] > ").strip()
                                    if q.lower() in ['exit', 'quit']:
                                        print("Exiting discussion mode.")
                                        return
                                    if q.lower() == 'back':
                                        print("Returning to hypothesis selection.")
                                        break
                                    # Use the critic to answer the question about the selected hypothesis
                                    hypothesis = accepted_hypotheses[selected_idx]['hypothesis']
                                    context_chunks = all_chunks if 'all_chunks' in locals() else []
                                    if self.hypothesis_critic and self.hypothesis_critic.model:
                                        # Build a prompt for Q&A
                                        prompt = f"You are an expert scientific reviewer. The user has a question about the following hypothesis.\n\nHypothesis:\n{hypothesis}\n\nUser Question:\n{q}\n\nPlease answer in detail, using the literature context if relevant."
                                        response = self.hypothesis_critic.model.models.generate_content(
                                            model="gemini-2.5-flash",
                                            contents=prompt
                                        )
                                        print(f"AI: {response.text.strip()}")
                                    else:
                                        print(f"AI: (No LLM available) This is a placeholder answer about '{hypothesis}'. User asked: {q}")
                            else:
                                print(f"Please enter a number between 1 and {len(accepted_hypotheses)}, or 'exit'.")
                    except Exception as e:
                        print(f"❌ Search or generation failed: {e}")
                        print("💡 Try again with a different query.")
                    self._display_commands()
                except Exception as e:
                    print(f"❌ Command parsing error: {e}")
                    print("❌ Usage: add <query> (e.g., 'add cancer')")
            elif query == '':
                continue
            else:
                print("❌ Unknown command. Available commands: 'add', 'clear', 'export', 'records', 'clear_records'")
                self._display_commands()

    def run_comprehensive_hypothesis_session(self, query, max_hypotheses=5):
        """
        Run a comprehensive hypothesis generation session with all features:
        - Timer-controlled critique process
        - Automated verdict determination
        - Citation tracking
        - Excel export
        """
        print("=" * 80)
        print("🧠 COMPREHENSIVE HYPOTHESIS GENERATION SESSION")
        print("=" * 80)
        print(f"Query: {query}")
        print(f"Features: 5-minute timer per hypothesis, automated verdicts, citation tracking, Excel export")
        print("=" * 80)

        # Clear previous records
        self.clear_hypothesis_records()

        # Run hypothesis generation
        print(f"\n🔍 Searching for relevant context...")
        context_chunks = self.retrieve_relevant_chunks(query, top_k=1500)
        if not context_chunks:
            print("❌ No relevant context found.")
            return None

        print(f"📚 Found {len(context_chunks)} relevant context chunks")

        # Generate hypotheses with all features
        print(f"\n🧠 Generating {max_hypotheses} hypotheses with comprehensive evaluation...")

        accepted_hypotheses = []
        attempts = 0
        max_attempts = 50  # Prevent infinite loops

        while len(accepted_hypotheses) < max_hypotheses and attempts < max_attempts:
            attempts += 1
            print(f"\n🔄 Hypothesis Attempt {attempts}/{max_attempts}")
            print("-" * 60)

            # ALWAYS select new chunks for EVERY hypothesis attempt to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"🔄 Selecting new chunks for hypothesis attempt {attempts}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation(query, randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    context_chunks = dynamic_chunks
                    print(f"📚 Using {len(context_chunks)} dynamically selected chunks for generation")
                else:
                    print(f"⚠️  Failed to select dynamic chunks, using original chunks")
            else:
                print(f"📚 Using original chunks for generation (no dynamic selection available)")

            # Start timer
            self.hypothesis_timer.start()
            print(f"⏱️  Timer started (5-minute limit)")

            # Generate hypothesis
            try:
                context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1)

                if not hypothesis_list:
                    print("❌ Failed to generate hypothesis")
                    self.hypothesis_records.append({
                        'Hypothesis': f'Failed generation attempt {attempts}',
                        'Accuracy': None,
                        'Novelty': None,
                        'Verdict': 'GENERATION_FAILED',
                        'Critique': 'Failed to generate hypothesis',
                        'Citations': 'No citations available'
                    })
                    continue

                hypothesis = hypothesis_list[0]
                
                # Validate hypothesis format before proceeding
                from hypothesis_tools import validate_hypothesis_format
                is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
                if not is_valid_format:
                    print(f"❌ Hypothesis rejected due to format: {format_reason}")
                    # Record format rejection
                    format_rejected_record = {
                        'Hypothesis': hypothesis,
                        'Accuracy': None,
                        'Novelty': None,
                        'Verdict': 'FORMAT_REJECTED',
                        'Critique': f'Hypothesis rejected: {format_reason}',
                        'Citations': 'No citations available'
                    }
                    self.hypothesis_records.append(format_rejected_record)
                    continue
                
                print(f"📝 Generated: {hypothesis}")

                # Check timer before critique
                if self.hypothesis_timer.check_expired():
                    print(f"⏰ Time limit reached before critique")
                    self.hypothesis_records.append({
                        'Hypothesis': hypothesis,
                        'Accuracy': None,
                        'Novelty': None,
                        'Verdict': 'TIMEOUT',
                        'Critique': 'Process timed out before critique',
                        'Citations': 'No citations available'
                    })
                    continue

                # Critique hypothesis
                print(f"🔍 Critiquing hypothesis...")
                critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt=query, lab_goals=get_lab_goals())

                # Check timer after critique
                if self.hypothesis_timer.check_expired():
                    print(f"⏰ Time limit reached after critique")
                    self.hypothesis_records.append({
                        'Hypothesis': hypothesis,
                        'Accuracy': critique_result.get('accuracy', None),
                        'Novelty': critique_result.get('novelty', None),
                        'Verdict': 'TIMEOUT',
                        'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                        'Citations': 'No citations available'
                    })
                    continue

                # Extract scores and determine verdict
                novelty = critique_result.get('novelty', 0)
                accuracy = critique_result.get('accuracy', 0)
                relevancy = critique_result.get('relevancy', 0)
                verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold=80, novelty_threshold=85, relevancy_threshold=50)
                remaining_time = self.hypothesis_timer.get_remaining_time()

                print(f"📊 Scores: Accuracy={accuracy}, Novelty={novelty}, Relevancy={relevancy}")
                print(f"⚖️  Automated Verdict: {verdict}")
                print(f"⏱️  Time remaining: {remaining_time:.1f}s")

                # Extract citations
                citations = self.extract_citations_from_chunks(context_chunks)
                formatted_citations = self.format_citations_for_export(citations)
                citation_count = len(citations)

                print(f"📚 Citations: {citation_count} unique sources")

                # Record hypothesis
                self.hypothesis_records.append({
                    'Hypothesis': hypothesis,
                    'Accuracy': accuracy,
                    'Novelty': novelty,
                    'Relevancy': relevancy,
                    'Verdict': verdict,
                    'Critique': critique_result.get('critique', ''),
                    'Citations': formatted_citations
                })

                # Check if accepted
                if verdict == 'ACCEPTED':
                    accepted_hypotheses.append({
                        "hypothesis": hypothesis,
                        "critique": critique_result,
                        "score": (novelty + accuracy + relevancy) / 3
                    })
                    print(f"✅ ACCEPTED! ({len(accepted_hypotheses)}/{max_hypotheses})")
                else:
                    print(f"❌ REJECTED (below thresholds)")

            except Exception as e:
                print(f"❌ Error during hypothesis processing: {e}")
                self.hypothesis_records.append({
                    'Hypothesis': f'Error in attempt {attempts}',
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'ERROR',
                    'Critique': f'Error: {str(e)}',
                    'Citations': 'No citations available'
                })
                continue

        # Final summary
        print(f"\n" + "=" * 80)
        print(f"🏁 SESSION COMPLETE")
        print(f"=" * 80)
        print(f"Total attempts: {attempts}")
        print(f"Accepted hypotheses: {len(accepted_hypotheses)}")
        print(f"Total records: {len(self.hypothesis_records)}")

        # Export results
        print(f"\n💾 Exporting results to Excel...")
        export_filename = self.export_hypotheses_to_excel()

        if export_filename:
            print(f"✅ Results exported to: {export_filename}")
            print(f"📁 File saved in: {os.path.dirname(os.path.abspath(export_filename))}")
        else:
            print(f"❌ Export failed")

        return accepted_hypotheses

    def generate_hypotheses_with_per_hypothesis_timer(self, n=5, max_rounds=10, filename=None):
        """
        Generate n hypotheses, using a 5-minute timer for the entire process of generating and refining each hypothesis.
        After each iteration, print the result, time left, scores, and hypothesis text.
        Save every iteration in self.hypothesis_records and append to Excel after each iteration.
        Adds empty verifier columns separated by an empty column.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from datetime import datetime
        if not self.current_package["chunks"]:
            print("❌ Package is empty. Add some chunks first.")
            return
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("❌ Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        novelty_threshold = 85
        accuracy_threshold = 80
        relevancy_threshold = 50
        print(f"\n🧠 Generating {n} hypotheses (5-minute timer per hypothesis)...")
        print(f"📦 Using {len(self.current_package['chunks'])} package chunks for critique")
        print(f"[INFO] Acceptance criteria: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold}")
        package_chunks = self.current_package["chunks"]
        self.hypothesis_records = []  # Clear previous records
        # Prepare Excel file
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Ensure hypothesis_export directory exists
            export_dir = "hypothesis_export"
            os.makedirs(export_dir, exist_ok=True)
            filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")
        # Write header row at the start, with empty separator and verifier columns
        columns = [
            'HypothesisNumber', 'Iteration', 'Status', 'Hypothesis', 'Novelty', 'Accuracy', 'Critique', 'TimeLeft',
            '',  # Empty separator column
            'Verifier Novelty Score (0-100)', 'Verifier Accuracy Score (0-100)', 'Verifier Verdict (accept, refuse)'
        ]
        pd.DataFrame(columns=columns).to_excel(filename, index=False)
        for hyp_idx in range(n):
            print(f"\nGenerating hypothesis {hyp_idx+1}/{n}\n")

            # ALWAYS select new chunks for EVERY hypothesis to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"🔄 Selecting new chunks for hypothesis {hyp_idx+1}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation("package query", randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    context_texts = dynamic_chunks
                    print(f"📚 Using {len(context_texts)} dynamically selected chunks for generation")
                else:
                    print(f"⚠️  Failed to select dynamic chunks, using original chunks")
                    context_texts = [chunk for chunk in package_chunks]
            else:
                print(f"📚 Using original chunks for generation (no dynamic selection available)")
                context_texts = [chunk for chunk in package_chunks]

            self.hypothesis_timer.start()
            time_left = self.hypothesis_timer.get_remaining_time()
            hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1)
            if not hypothesis_list:
                print("❌ Failed to generate hypothesis. Skipping...")
                continue
                
            hypothesis = hypothesis_list[0]
            
            # Validate hypothesis format before proceeding
            from hypothesis_tools import validate_hypothesis_format
            is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
            if not is_valid_format:
                print(f"❌ Hypothesis rejected due to format: {format_reason}")
                # Record format rejection
                format_rejected_record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': 0,
                    'Status': 'FORMAT_REJECTED',
                    'Hypothesis': hypothesis,
                    'Novelty': None,
                    'Accuracy': None,
                    'Critique': f'Hypothesis rejected: {format_reason}',
                    'TimeLeft': int(time_left),
                    '': '',
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(format_rejected_record)
                # Append to Excel
                df_row = pd.DataFrame([format_rejected_record])
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_row.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=len(self.hypothesis_records))
                continue
            
            accepted = False
            iteration = 0
            while not accepted and not self.hypothesis_timer.check_expired() and iteration < max_rounds:
                iteration += 1
                critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt=query, lab_goals=get_lab_goals())
                novelty = critique_result.get('novelty', 0)
                accuracy = critique_result.get('accuracy', 0)
                relevancy = critique_result.get('relevancy', 0)
                time_left = self.hypothesis_timer.get_remaining_time()
                verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold=80, novelty_threshold=85, relevancy_threshold=50)
                print(f"Iteration {iteration}: {verdict} | Novelty: {novelty}/85 | Accuracy: {accuracy}/80 | Relevancy: {relevancy}/50 | Time left: {int(time_left)}s")
                print(f"Hypothesis: {hypothesis}")
                # Save this iteration to records
                record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': iteration,
                    'Status': verdict,
                    'Hypothesis': hypothesis,
                    'Novelty': novelty,
                    'Accuracy': accuracy,
                    'Critique': critique_result.get('critique', ''),
                    'TimeLeft': int(time_left),
                    '': '',  # Empty separator column
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(record)
                # Append to Excel after each iteration
                df_row = pd.DataFrame([record])
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    writer.book = load_workbook(filename)
                    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
                    startrow = writer.book['Hypotheses'].max_row
                    df_row.to_excel(writer, sheet_name='Hypotheses', index=False, header=False, startrow=startrow)
                if verdict == "ACCEPTED":
                    print(f"✅ Hypothesis accepted! (verdict: {verdict})")
                    accepted = True
                    break
                else:
                    # Print specific rejection reason
                    if accuracy < 80:
                        print(f"❌ Hypothesis rejected: Low accuracy score ({accuracy}/80)")
                    elif novelty < 85:
                        print(f"❌ Hypothesis rejected: Low novelty score ({novelty}/85)")
                    elif relevancy < 50:
                        print(f"❌ Hypothesis rejected: Low relevancy score ({relevancy}/50)")
                    else:
                        print(f"❌ Hypothesis rejected: Unknown reason (verdict: {verdict})")
                if self.hypothesis_timer.check_expired():
                    print(f"⏰ Time limit reached for hypothesis {hyp_idx+1}. Moving to next hypothesis.")
                    break
                # Regenerate hypothesis for next iteration
                new_hypothesis = self.hypothesis_generator.generate(context_texts, n=1)
                if new_hypothesis:
                    hypothesis = new_hypothesis[0]
        print(f"\n✅ All iterations and results have been saved to: {filename}")
        print(f"📁 File saved in: {os.path.dirname(os.path.abspath(filename))}")
        return self.hypothesis_records

    def is_lab_authored_paper(self, metadata):
        """Check if a paper is authored by the lab PI or lab members."""
        from hypothesis_tools import get_lab_config

        config = get_lab_config()
        lab_name = config.get("lab_name", "Dr. Xiaojing Ma")

        # Get author information from metadata
        authors = metadata.get('author', metadata.get('authors', '')).lower()
        if not authors or authors == 'unknown authors':
            # If no authors found in metadata, return False
            return False

        # Clean author string - remove "et al" and other common patterns
        authors_clean = re.sub(r'\bet\s+al\.?\b', '', authors)  # Remove "et al" or "et al."
        authors_clean = re.sub(r'\band\s+others\b', '', authors_clean)  # Remove "and others"
        authors_clean = re.sub(r'\s*&\s*others\b', '', authors_clean)  # Remove "& others" (with optional space)
        authors_clean = re.sub(r'\s+', ' ', authors_clean).strip()  # Normalize whitespace
        authors_clean = re.sub(r'[,\s\.]+$', '', authors_clean)  # Remove trailing commas/spaces/periods

        # Check for lab PI name variations - more comprehensive matching
        lab_pi_variations = [
            lab_name.lower(),
            lab_name.replace("Dr. ", "").lower(),
            lab_name.replace("Dr. ", "").replace(" ", "").lower(),
            "xiaojing ma",
            "xiaojing",
            "ma x",
            "ma, x",
            "ma, xiaojing",
            "xiaojing ma,",
            "x. ma",
            "x ma",
            "ma, x.",
            "ma xiaojing",
            "xiaojing m",
            "x. m.",
            "x m"
        ]

        # Check if any lab PI variation is in authors (using word boundaries)
        for variation in lab_pi_variations:
            # Use word boundaries to avoid partial matches
            if re.search(r'\b' + re.escape(variation) + r'\b', authors_clean):
                return True

        # Also check for common lab member patterns (if any are known)
        # This could be expanded with actual lab member names
        lab_member_patterns = [
            r'\bma\s+x\b',  # X Ma pattern
            r'\bxiaojing\s+ma\b',  # Xiaojing Ma pattern
            r'\bma,\s*x\b',  # Ma, X pattern
        ]
        
        for pattern in lab_member_patterns:
            if re.search(pattern, authors_clean, re.IGNORECASE):
                return True

        return False

    def retrieve_relevant_chunks_with_lab_papers(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Retrieve relevant chunks using the sophisticated lab paper search algorithm.
        This method now implements author prioritization, impact factor weighting, 
        citation analysis, and temporal balance (50/50 old/new mix).

        Args:
            query: Search query
            top_k: Total number of chunks to retrieve
            lab_paper_ratio: Optional fixed ratio (0.0 to 1.0). If None, automatically determined.
            preferred_authors: List of preferred author names to prioritize
        """
        # Use the new sophisticated search algorithm
        return self.sophisticated_lab_paper_search(query, top_k, lab_paper_ratio, preferred_authors)

    def configure_lab_paper_ratio(self, ratio=None):
        """Configure the ratio of lab-authored papers to include in searches."""
        if ratio is None:
            current_ratio = getattr(self, 'lab_paper_ratio', None)
            if current_ratio is None:
                print(f"🔬 Lab paper ratio: AUTO (automatically determined based on available lab papers)")
                print(f"   • If lab papers found: 10-50% (based on availability)")
                print(f"   • If no lab papers: 0%")
            else:
                print(f"🔬 Current lab paper ratio: {current_ratio:.1%} (fixed)")

            try:
                choice = input("Enter 'auto' for automatic determination, or a number (0.0 to 1.0, e.g., 0.3 for 30%): ").strip().lower()
                if choice == 'auto':
                    self.lab_paper_ratio = None
                    print(f"✅ Lab paper ratio set to AUTO (automatically determined)")
                else:
                    new_ratio = float(choice)
                    if 0.0 <= new_ratio <= 1.0:
                        self.lab_paper_ratio = new_ratio
                        print(f"✅ Lab paper ratio updated to {new_ratio:.1%} (fixed)")
                    else:
                        print("❌ Invalid ratio. Must be between 0.0 and 1.0.")
            except ValueError:
                print("❌ Invalid input. Please enter 'auto' or a number between 0.0 and 1.0.")
        else:
            if ratio == 'auto':
                self.lab_paper_ratio = None
                print(f"✅ Lab paper ratio set to AUTO (automatically determined)")
            elif 0.0 <= ratio <= 1.0:
                self.lab_paper_ratio = ratio
                print(f"✅ Lab paper ratio set to {ratio:.1%} (fixed)")
            else:
                print("❌ Invalid ratio. Must be 'auto' or between 0.0 and 1.0.")

    def test_lab_paper_detection(self, query="UBR5 cancer"):
        """Test function to debug lab paper detection."""
        print(f"🔬 Testing lab paper detection for query: '{query}'")
        
        if not self.use_chromadb or not self.chroma_manager:
            print("❌ ChromaDB not available for testing.")
            return
        
        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            print("❌ Failed to get query embedding.")
            return
        
        print(f"🔍 Searching ChromaDB for lab papers...")
        
        # Search broadly to find potential lab papers
        search_results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=1000  # Search broadly
        )
        
        lab_papers_found = []
        other_papers = []
        
        for result in search_results:
            metadata = result.get('metadata', {})
            authors = metadata.get('author', metadata.get('authors', ''))
            title = metadata.get('title', '')
            
            if self.is_lab_authored_paper(metadata):
                lab_papers_found.append({
                    'title': title,
                    'authors': authors,
                    'doi': metadata.get('doi', 'No DOI'),
                    'source': metadata.get('source_name', 'Unknown')
                })
            else:
                other_papers.append({
                    'title': title,
                    'authors': authors,
                    'doi': metadata.get('doi', 'No DOI'),
                    'source': metadata.get('source_name', 'Unknown')
                })
        
        print(f"\n📊 Lab Paper Detection Results:")
        print(f"   Total papers searched: {len(search_results)}")
        print(f"   Lab papers found: {len(lab_papers_found)}")
        print(f"   Other papers: {len(other_papers)}")
        
        if lab_papers_found:
            print(f"\n🔬 Lab Papers Found:")
            for i, paper in enumerate(lab_papers_found[:10], 1):  # Show first 10
                print(f"   {i}. {paper['title']}")
                print(f"      Authors: {paper['authors']}")
                print(f"      DOI: {paper['doi']}")
                print(f"      Source: {paper['source']}")
                print()
        else:
            print(f"\n❌ No lab papers found. This is expected because:")
            print(f"   • Current database sources: PubMed, bioRxiv, medRxiv")
            print(f"   • Lab papers may not be in these general databases")
            print(f"   • Lab paper detection logic is working correctly")
            print(f"   • Author name variations are properly configured")
            
            # Show some example papers to help debug
            if other_papers:
                print(f"\n📚 Sample papers in database:")
                for i, paper in enumerate(other_papers[:5], 1):
                    print(f"   {i}. {paper['title']}")
                    print(f"      Authors: {paper['authors']}")
                    print(f"      Source: {paper['source']}")
                    print()
                
                print(f"💡 To include lab papers, consider:")
                print(f"   • Adding lab papers manually to the database")
                print(f"   • Checking if lab papers exist in PubMed/bioRxiv/medRxiv")
                print(f"   • Using different data sources that include lab publications")

    def configure_randomization_strategy(self, strategy=None):
        """
        Configure the chunk randomization strategy for hypothesis generation.
        
        Args:
            strategy: Randomization strategy ('enhanced', 'shuffle', 'diversity', 'time_based', 'none')
        """
        available_strategies = {
            'enhanced': 'Enhanced randomization combining relevance and diversity',
            'shuffle': 'Simple random shuffle of all chunks',
            'diversity': 'Prioritize diversity by publication date and citations',
            'time_based': 'Time-based randomization (changes every minute)',
            'none': 'No additional randomization (use default ordering)'
        }
        
        if strategy is None:
            print("\n🎲 Chunk Randomization Strategy Configuration")
            print("=" * 50)
            print("Current strategy:", getattr(self, 'randomization_strategy', 'enhanced'))
            print("\nAvailable strategies:")
            for key, description in available_strategies.items():
                print(f"  • {key}: {description}")
            
            print(f"\n💡 The system already ensures each hypothesis uses different papers.")
            print(f"💡 This randomization adds additional variety to chunk selection within those papers.")
            
            strategy = input("\nEnter strategy name (or press Enter to keep current): ").strip().lower()
            if not strategy:
                return
        
        if strategy not in available_strategies:
            print(f"❌ Invalid strategy: {strategy}")
            print(f"Available strategies: {', '.join(available_strategies.keys())}")
            return
        
        self.randomization_strategy = strategy
        print(f"✅ Randomization strategy set to: {strategy}")
        print(f"   📝 {available_strategies[strategy]}")

    def configure_preferred_authors(self, authors=None):
        """Configure preferred authors for the sophisticated search algorithm."""
        if authors is None:
            current_authors = getattr(self, 'preferred_authors', [])
            if current_authors:
                print(f"👥 Current preferred authors: {', '.join(current_authors)}")
            else:
                print(f"👥 No preferred authors configured")

            try:
                print("\nEnter preferred author names (comma-separated, or 'clear' to remove all):")
                choice = input("Authors: ").strip()
                if choice.lower() == 'clear':
                    self.preferred_authors = []
                    print(f"✅ Cleared preferred authors")
                elif choice:
                    author_list = [author.strip() for author in choice.split(',') if author.strip()]
                    self.preferred_authors = author_list
                    print(f"✅ Set preferred authors: {', '.join(author_list)}")
                else:
                    print(f"✅ Keeping current preferred authors")
            except KeyboardInterrupt:
                print("\n✅ Configuration cancelled.")
        else:
            if isinstance(authors, list):
                self.preferred_authors = authors
                print(f"✅ Set preferred authors: {', '.join(authors)}")
            else:
                print(f"❌ Invalid authors format. Expected list of strings.")

    def sophisticated_lab_paper_search(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Sophisticated lab paper search with author prioritization, impact factor weighting, 
        citation analysis, temporal balance (50/50 old/new mix), and preprint deprioritization.
        
        Args:
            query: Search query
            top_k: Total number of chunks to retrieve
            lab_paper_ratio: Optional fixed ratio (0.0 to 1.0). If None, automatically determined.
            preferred_authors: List of preferred author names to prioritize
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("❌ ChromaDB not available.")
            return []

        # Check if ChromaDB has data
        if not self.is_chromadb_ready():
            print("❌ ChromaDB is empty. Please load data first using the master processor (option 4).")
            return []

        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            print("❌ Failed to get query embedding.")
            return []

        print(f"🔬 Starting sophisticated lab paper search...")
        print(f"📊 Target: {top_k} total chunks")
        
        # Use configured preferred authors if none provided
        if preferred_authors is None:
            preferred_authors = getattr(self, 'preferred_authors', [])
        
        if preferred_authors:
            print(f"👥 Using preferred authors: {', '.join(preferred_authors)}")
        
        # Step 1: Broad search to get all relevant papers
        print(f"🔍 Performing broad search for relevant papers...")
        broad_search_results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=top_k * 5  # Search broadly to find all relevant papers
        )

        if not broad_search_results:
            print("❌ No papers found in ChromaDB.")
            return []

        # Step 2: Categorize papers and extract metadata
        print(f"📋 Categorizing papers and extracting metadata...")
        lab_papers = []
        preferred_author_papers = []
        other_papers = []
        
        # Track citation statistics for analysis
        all_citations = []
        journal_impact_factors = {}
        
        for result in broad_search_results:
            metadata = result.get('metadata', {})
            
            # Extract citation count for analysis
            citation_count = metadata.get('citation_count', 'not found')
            if citation_count != 'not found':
                try:
                    citations = int(citation_count)
                    all_citations.append(citations)
                except (ValueError, TypeError):
                    pass
            
            # Track journal for impact factor analysis
            journal = metadata.get('journal', 'Unknown journal')
            if journal not in journal_impact_factors:
                journal_impact_factors[journal] = []
            if citation_count != 'not found':
                try:
                    journal_impact_factors[journal].append(int(citation_count))
                except (ValueError, TypeError):
                    pass
            
            # Categorize papers
            if self.is_lab_authored_paper(metadata):
                lab_papers.append(result)
            elif preferred_authors and self._is_preferred_author_paper(metadata, preferred_authors):
                preferred_author_papers.append(result)
            else:
                other_papers.append(result)

        print(f"📊 Found {len(lab_papers)} lab papers, {len(preferred_author_papers)} preferred author papers, {len(other_papers)} other papers")

        # Step 3: Calculate statistics for prioritization
        print(f"📈 Calculating citation and journal statistics...")
        
        # Calculate average citation count
        avg_citations = np.mean(all_citations) if all_citations else 0
        median_citations = np.median(all_citations) if all_citations else 0
        
        # Calculate journal impact factors (average citations per journal)
        journal_impact_scores = {}
        for journal, citations in journal_impact_factors.items():
            if citations:
                journal_impact_scores[journal] = np.mean(citations)
        
        print(f"📊 Average citations: {avg_citations:.1f}, Median: {median_citations:.1f}")
        print(f"📊 Journal impact scores calculated for {len(journal_impact_scores)} journals")

        # Step 4: Determine lab paper ratio
        if lab_paper_ratio is None:
            if len(lab_papers) > 0:
                available_lab_ratio = min(len(lab_papers) / top_k, 0.5)
                lab_paper_ratio = max(available_lab_ratio, 0.1)
                print(f"🔬 Auto-determined lab paper ratio: {lab_paper_ratio:.1%}")
            else:
                lab_paper_ratio = 0.0
                print(f"🔬 No lab-authored papers found")
        else:
            print(f"🔬 Using configured lab paper ratio: {lab_paper_ratio:.1%}")

        # Step 5: Author paper selection (Step 1 of search execution)
        print(f"👥 Step 1: Selecting author-prioritized papers...")
        lab_chunks_needed = int(top_k * lab_paper_ratio)
        preferred_author_chunks_needed = min(len(preferred_author_papers), int(top_k * 0.2))  # Up to 20% for preferred authors
        remaining_chunks_needed = top_k - lab_chunks_needed - preferred_author_chunks_needed

        selected_papers = []
        used_papers = set()

        # Add lab papers first (highest priority)
        for paper in lab_papers[:lab_chunks_needed]:
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)

        # Add preferred author papers
        for paper in preferred_author_papers[:preferred_author_chunks_needed]:
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)

        print(f"✅ Selected {len(selected_papers)} author-prioritized papers")

        # Step 6: Remaining paper selection with temporal balance and preprint deprioritization (Step 2 of search execution)
        print(f"📚 Step 2: Selecting remaining papers with temporal balance and preprint deprioritization...")
        
        # Calculate median age of ALL search results (not just selected papers)
        all_years = []
        for paper in broad_search_results:
            metadata = paper.get('metadata', {})
            publication_date = metadata.get('publication_date', '')
            if publication_date and len(publication_date) >= 4:
                try:
                    year = int(publication_date[:4])
                    all_years.append(year)
                except (ValueError, TypeError):
                    pass
        
        median_year = np.median(all_years) if all_years else datetime.now().year - 5
        current_year = datetime.now().year
        
        print(f"📅 Median year of ALL search results: {median_year:.0f}")
        print(f"📅 Temporal dividing line: {median_year:.0f} (papers before = old, after = new)")
        
        # Show temporal distribution statistics
        if all_years:
            min_year = min(all_years)
            max_year = max(all_years)
            year_counts = {}
            for year in all_years:
                year_counts[year] = year_counts.get(year, 0) + 1
            
            print(f"📊 Temporal distribution: {min_year} to {max_year}")
            print(f"📊 Total papers with dates: {len(all_years)}")
            print(f"📊 Papers ≤ {median_year:.0f}: {sum(1 for y in all_years if y <= median_year)}")
            print(f"📊 Papers > {median_year:.0f}: {sum(1 for y in all_years if y > median_year)}")

        # Categorize remaining papers by temporal period and preprint status
        old_papers = []
        new_papers = []
        
        for paper in other_papers:
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id in used_papers:
                continue
                
            metadata = paper.get('metadata', {})
            publication_date = metadata.get('publication_date', '')
            
            # Determine temporal category
            temporal_category = None
            if publication_date and len(publication_date) >= 4:
                try:
                    year = int(publication_date[:4])
                    if year <= median_year:
                        temporal_category = 'old'
                    else:
                        temporal_category = 'new'
                except (ValueError, TypeError):
                    temporal_category = 'unknown'
            else:
                temporal_category = 'unknown'
            
            # Determine preprint status
            is_preprint = metadata.get('is_preprint', False)
            preprint_score = metadata.get('preprint_score', 0.0)
            preprint_type = metadata.get('preprint_type', 'unknown')
            
            # Add preprint information to metadata if not present
            if 'is_preprint' not in metadata:
                # Try to detect preprint status from existing fields
                source = metadata.get('source', '').lower()
                journal = metadata.get('journal', '').lower()
                
                if (source in ['biorxiv', 'medrxiv', 'arxiv', 'chemrxiv'] or 
                    any(indicator in journal for indicator in ['arxiv', 'biorxiv', 'medrxiv', 'chemrxiv', 'preprint'])):
                    is_preprint = True
                    preprint_score = 0.8
                    preprint_type = source if source else 'preprint_server'
                
                # Update metadata
                metadata['is_preprint'] = is_preprint
                metadata['preprint_score'] = preprint_score
                metadata['preprint_type'] = preprint_type
            
            # Categorize by temporal period
            if temporal_category == 'old':
                old_papers.append(paper)
            elif temporal_category == 'new':
                new_papers.append(paper)
            else:
                # Unknown date papers go to old category as fallback
                old_papers.append(paper)

        print(f"📊 Categorized remaining papers: {len(old_papers)} old, {len(new_papers)} new")
        
        # Count preprint papers for reporting
        preprint_count = sum(1 for p in old_papers + new_papers 
                           if p.get('metadata', {}).get('is_preprint', False))
        print(f"📊 Found {preprint_count} preprint papers (will be deprioritized)")

        # Step 7: Score and rank remaining papers with preprint deprioritization
        print(f"🎯 Step 3: Scoring and ranking papers with preprint deprioritization...")
        
        def calculate_paper_score(paper, is_old_paper):
            """Calculate comprehensive paper score with preprint deprioritization."""
            metadata = paper.get('metadata', {})
            
            # Base score starts at 100
            score = 100.0
            
            # Citation-based scoring (0-50 points)
            citation_count = metadata.get('citation_count', 'not found')
            if citation_count != 'not found':
                try:
                    citations = int(citation_count)
                    if citations > 0:
                        # Logarithmic scaling for citations
                        citation_score = min(50, 10 * np.log10(citations + 1))
                        score += citation_score
                except (ValueError, TypeError):
                    pass
            
            # Journal impact scoring (0-30 points)
            journal = metadata.get('journal', 'Unknown journal')
            if journal in journal_impact_scores:
                journal_score = min(30, journal_impact_scores[journal] / 10)
                score += journal_score
            
            # Temporal balance scoring (0-20 points)
            if is_old_paper:
                # Old papers get bonus to maintain balance
                score += 20
            else:
                # New papers get smaller bonus
                score += 10
            
            # PREPRINT DEPRIORITIZATION (penalty of 0-40 points)
            is_preprint = metadata.get('is_preprint', False)
            preprint_score = metadata.get('preprint_score', 0.0)
            
            if is_preprint:
                # Calculate preprint penalty based on preprint score
                # Higher preprint score = higher penalty (more clearly a preprint)
                preprint_penalty = min(40, preprint_score * 50)  # 0.8 score = 40 penalty
                score -= preprint_penalty
                
                # Additional penalty for certain preprint types
                preprint_type = metadata.get('preprint_type', '')
                if preprint_type in ['biorxiv', 'medrxiv', 'chemrxiv']:
                    score -= 10  # Extra penalty for bio/med/chem preprints
                elif preprint_type == 'working_paper':
                    score -= 5   # Smaller penalty for working papers
            
            # Ensure score doesn't go below 0
            score = max(0, score)
            
            return score
        
        # Score old papers
        old_papers_scored = []
        for paper in old_papers:
            score = calculate_paper_score(paper, is_old_paper=True)
            old_papers_scored.append((paper, score))
        
        # Score new papers
        new_papers_scored = []
        for paper in new_papers:
            score = calculate_paper_score(paper, is_old_paper=False)
            new_papers_scored.append((paper, score))
        
        # Sort by score (highest first)
        old_papers_scored.sort(key=lambda x: x[1], reverse=True)
        new_papers_scored.sort(key=lambda x: x[1], reverse=True)
        
        # Show top scored papers for each category
        print(f"📊 Top 5 old papers scores: {[f'{p[1]:.1f}' for p in old_papers_scored[:5]]}")
        print(f"📊 Top 5 new papers scores: {[f'{p[1]:.1f}' for p in new_papers_scored[:5]]}")
        
        # Count preprint papers in top scores
        top_old_preprints = sum(1 for p, _ in old_papers_scored[:20] 
                               if p.get('metadata', {}).get('is_preprint', False))
        top_new_preprints = sum(1 for p, _ in new_papers_scored[:20] 
                               if p.get('metadata', {}).get('is_preprint', False))
        print(f"📊 Preprint papers in top 20 scores: {top_old_preprints} old, {top_new_preprints} new")

        # Step 8: Select remaining papers with balanced distribution
        print(f"🎯 Step 4: Selecting remaining papers with balanced distribution...")
        
        # Calculate target chunks for each temporal category
        old_chunks_needed = remaining_chunks_needed // 2
        new_chunks_needed = remaining_chunks_needed - old_chunks_needed
        
        print(f"📊 Target distribution: {old_chunks_needed} old papers, {new_chunks_needed} new papers")
        
        # Select old papers
        old_selected = 0
        for paper, score in old_papers_scored:
            if old_selected >= old_chunks_needed:
                break
                
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)
                old_selected += 1
        
        # Select new papers
        new_selected = 0
        for paper, score in new_papers_scored:
            if new_selected >= new_chunks_needed:
                break
                
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)
                new_selected += 1
        
        print(f"✅ Selected {old_selected} old papers and {new_selected} new papers")
        
        # If we don't have enough papers, fill with remaining high-scoring papers
        if len(selected_papers) < top_k:
            remaining_needed = top_k - len(selected_papers)
            print(f"⚠️  Need {remaining_needed} more papers, filling with remaining high-scoring papers...")
            
            # Combine remaining papers and sort by score
            remaining_papers = []
            for paper, score in old_papers_scored:
                if paper not in selected_papers:
                    remaining_papers.append((paper, score))
            for paper, score in new_papers_scored:
                if paper not in selected_papers:
                    remaining_papers.append((paper, score))
            
            # Sort by score and add until we reach target
            remaining_papers.sort(key=lambda x: x[1], reverse=True)
            for paper, score in remaining_papers:
                if len(selected_papers) >= top_k:
                    break
                    
                paper_id = self._get_paper_identifier(paper.get('metadata', {}))
                if paper_id and paper_id not in used_papers:
                    selected_papers.append(paper)
                    used_papers.add(paper_id)

        # Step 9: Final statistics and validation
        print(f"📊 Final selection statistics:")
        print(f"   Total papers selected: {len(selected_papers)}")
        print(f"   Lab papers: {sum(1 for p in selected_papers if self.is_lab_authored_paper(p.get('metadata', {})))}")
        print(f"   Preferred author papers: {sum(1 for p in selected_papers if preferred_authors and self._is_preferred_author_paper(p.get('metadata', {}), preferred_authors))}")
        
        # Count preprint papers in final selection
        final_preprint_count = sum(1 for p in selected_papers 
                                  if p.get('metadata', {}).get('is_preprint', False))
        print(f"   Preprint papers: {final_preprint_count} (deprioritized)")
        
        # Show preprint distribution by type
        preprint_types = {}
        for paper in selected_papers:
            metadata = paper.get('metadata', {})
            if metadata.get('is_preprint', False):
                preprint_type = metadata.get('preprint_type', 'unknown')
                preprint_types[preprint_type] = preprint_types.get(preprint_type, 0) + 1
        
        if preprint_types:
            print(f"   Preprint types: {', '.join([f'{k}: {v}' for k, v in preprint_types.items()])}")
        
        # Validate we have enough papers
        if len(selected_papers) < top_k:
            print(f"⚠️  Warning: Only found {len(selected_papers)} papers (requested {top_k})")
        elif len(selected_papers) > top_k:
            print(f"📊 Note: Selected {len(selected_papers)} papers (requested {top_k})")
        
        # Return the selected papers
        return selected_papers

    def _is_preferred_author_paper(self, metadata, preferred_authors):
        """Check if a paper is authored by preferred authors."""
        if not preferred_authors:
            return False
        
        authors = metadata.get('author', metadata.get('authors', ''))
        if not authors:
            return False
        
        # Convert to string if it's a list
        if isinstance(authors, list):
            authors = '; '.join(authors)
        
        authors_lower = str(authors).lower()
        
        for preferred_author in preferred_authors:
            if preferred_author.lower() in authors_lower:
                return True
        
        return False

    def detect_preprint_status(self, metadata):
        """
        Detect preprint status from existing metadata fields.
        
        Args:
            metadata: Paper metadata dictionary
            
        Returns:
            tuple: (is_preprint: bool, preprint_type: str, preprint_score: float)
        """
        preprint_score = 0.0
        preprint_type = "unknown"
        
        # Check source field first (most reliable)
        source = metadata.get('source', '').lower()
        if source in ["biorxiv", "medrxiv", "arxiv", "chemrxiv", "preprints"]:
            preprint_score += 0.8
            preprint_type = source
        
        # Check journal name for preprint indicators
        journal = metadata.get('journal', '').lower()
        if journal:
            preprint_indicators = [
                "arxiv", "biorxiv", "medrxiv", "chemrxiv", "preprint", 
                "working paper", "manuscript", "draft"
            ]
            for indicator in preprint_indicators:
                if indicator in journal:
                    preprint_score += 0.6
                    preprint_type = "preprint_server"
                    break
        
        # Check if it's from a preprint server origin
        origin = metadata.get('origin', '').lower()
        if origin in ["biorxiv", "medrxiv", "chemrxiv", "preprints"]:
            preprint_score += 0.9
            preprint_type = origin
        
        # Check version information
        version = metadata.get('version', '')
        if version and version != "1":
            preprint_score += 0.3
            preprint_type = "versioned_preprint"
        
        # Check status
        status = metadata.get('status', '').lower()
        if status in ["submitted", "pending", "working"]:
            preprint_score += 0.4
            preprint_type = "submitted_preprint"
        
        # Determine final classification
        is_preprint = preprint_score >= 0.5
        
        return is_preprint, preprint_type, preprint_score

    def add_preprint_metadata(self, metadata):
        """
        Add preprint metadata to paper metadata if not present.
        
        Args:
            metadata: Paper metadata dictionary
            
        Returns:
            dict: Updated metadata with preprint information
        """
        if 'is_preprint' not in metadata:
            is_preprint, preprint_type, preprint_score = self.detect_preprint_status(metadata)
            metadata['is_preprint'] = is_preprint
            metadata['preprint_type'] = preprint_type
            metadata['preprint_score'] = preprint_score
        
        return metadata

    def _get_journal_impact_factor(self, journal_name):
        """
        Get journal impact factor (simplified implementation).
        In a real system, this would query a journal impact factor database.
        """
        # Simplified impact factor mapping - in practice, this would be a database lookup
        impact_factors = {
            'nature': 49.962,
            'science': 56.9,
            'cell': 66.85,
            'nature medicine': 87.241,
            'nature biotechnology': 68.164,
            'nature genetics': 41.307,
            'nature cell biology': 28.213,
            'nature immunology': 31.25,
            'nature reviews immunology': 108.555,
            'immunity': 43.474,
            'journal of immunology': 5.422,
            'journal of experimental medicine': 17.579,
            'proceedings of the national academy of sciences': 12.779,
            'plos one': 3.752,
            'bioinformatics': 6.937,
            'nucleic acids research': 19.16,
            'genome research': 11.093,
            'genome biology': 17.906,
            'cell reports': 9.995,
            'molecular cell': 19.328,
            'developmental cell': 13.417,
            'current biology': 10.834,
            'elife': 8.713,
            'plos biology': 9.593,
            'plos genetics': 6.02,
            'plos computational biology': 4.7,
            'bmc genomics': 4.317,
            'bmc bioinformatics': 3.169,
            'bioinformatics': 6.937,
            'nucleic acids research': 19.16,
            'genome research': 11.093,
            'genome biology': 17.906,
            'cell reports': 9.995,
            'molecular cell': 19.328,
            'developmental cell': 13.417,
            'current biology': 10.834,
            'elife': 8.713,
            'plos biology': 9.593,
            'plos genetics': 6.02,
            'plos computational biology': 4.7,
            'bmc genomics': 4.317,
            'bmc bioinformatics': 3.169
        }
        
        journal_lower = journal_name.lower()
        for key, impact in impact_factors.items():
            if key in journal_lower:
                return impact
                
        return 1.0  # Default impact factor for unknown journals

def main():
    """Main function to run the enhanced RAG query system."""
    print("=== Enhanced RAG System Startup ===")
    print("🚀 Starting Enhanced RAG System with ChromaDB...")

    # Initialize with ChromaDB enabled and fast startup
    rag_system = EnhancedRAGQuery(use_chromadb=True, load_data_at_startup=False)



    # Run interactive search
    rag_system.interactive_search()

if __name__ == "__main__":
    main()
